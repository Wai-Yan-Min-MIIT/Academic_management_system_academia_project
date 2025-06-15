# academia/views.py

import pandas as pd
import os
from datetime import date
import datetime
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.sessions.models import Session
from django.contrib.auth.decorators import login_required
from django.contrib.auth.signals import user_logged_in
from django.conf import settings
from .auth_backends import MIITUserBackend
from django.http import HttpResponseForbidden, HttpResponse
from django.http import JsonResponse
from django.db import connection
from collections import defaultdict
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Subquery
from django.contrib import messages
from django.dispatch import receiver
from django.core.mail import send_mass_mail, send_mail
from academia.models import *
from academia.forms import *
from academia.utils import fetch_news_articles
from django.views.decorators.csrf import csrf_exempt
# from django.templatetags.static import static
from django.db.models import Q
from openpyxl import load_workbook
from django.db import transaction
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from io import BytesIO
import os
from django.conf import settings
import json
import bcrypt
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.utils.crypto import get_random_string
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.urls import reverse
import requests
import qrcode
import base64
from io import BytesIO
from PIL import Image,ImageDraw
import pyotp


def user_login(request):
    if request.method == 'POST':
        email = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=email, password=password)

        if user and user.Deleted == 0 and not user.acc_locked:
            if user.mfa_enabled:
                request.session['pre_mfa_user_id'] = user.UserID
                return redirect('mfa_verify')
            else:
                login(request, user)
                return redirect('choose_role')
                
        elif user and user.Deleted == 1 and not user.acc_locked:
            messages.error(request, 'User account does not exist!', extra_tags='no_account')

        elif user and user.acc_locked:
            messages.info(request, 'Your account is locked!', extra_tags='account_locked')

        else:
            messages.error(request, 'Invalid username or password.', extra_tags='incorrect')

    return render(request, 'login.html')


def mfa_verify(request):
    if request.method == 'POST':
        user_id = request.session.get('pre_mfa_user_id')
        if not user_id:
            messages.error(request, 'Session expired. Please log in again.', extra_tags='session_expired')
            return redirect('user_login')

        user = MIITUsers.objects.get(UserID=user_id)

        if 'token' in request.POST:
            otp = request.POST.get('token')
            totp = pyotp.TOTP(user.mfa_secret)
            if totp.verify(otp):
                login(request, user)
                request.session.pop('pre_mfa_user_id', None)  # Clear the session key after successful login
                return redirect('choose_role')
            else:
                messages.error(request, 'Invalid OTP.', extra_tags='invalid_code')
        elif 'backup_code' in request.POST:
            backup_code = request.POST.get('backup_code')
            if backup_code in user.backup_codes.split(','):
                user.backup_codes = ','.join(code for code in user.backup_codes.split(',') if code != backup_code)
                user.save()
                login(request, user)
                request.session.pop('pre_mfa_user_id', None)  # Clear the session key after successful login
                return redirect('choose_role')
            else:
                messages.error(request, 'Invalid backup code.', extra_tags='invalid_code')

    return render(request, 'verify_mfa.html')



def user_logout(request):
    UserDevice.objects.filter(session_key=request.session.session_key).delete()
    logout(request)
    request.session.flush()
    messages.info(request, 'You have logged out!', extra_tags='logged_out')
    return redirect('login')


@login_required
def logout_device(request, session_key):
    # Get the current session key
    current_session_key = request.session.session_key

    # Prevent deletion of the current session
    if current_session_key == session_key:
        messages.error(request, "You cannot log out the device you are currently using.", extra_tags='cannot_logout_current_device')
        return HttpResponseForbidden("You cannot log out the device you are currently using.")

    # Find the session and delete it
    session = get_object_or_404(Session, session_key=session_key)
    session.delete()

    # Also delete the corresponding UserDevice entry
    UserDevice.objects.filter(session_key=session_key).delete()

    messages.success(request, "Device logged out successfully.", extra_tags='device_logged_out')
    return redirect('session_management')


@login_required
def session_management(request):
    user = request.user
    devices = UserDevice.objects.filter(user=user)
    context = {
        'devices': devices
    }
    return render(request, 'session_management.html', context)



@login_required
def choose_role(request):
    user_roles = MIITRole.objects.filter(RoleID__in=MIITUserRole.objects.filter(UserID=request.user.UserID).values_list('RoleID', flat=True))
    if user_roles and len(user_roles) > 1:
        user_roles = request.session.get('user_roles', [])
        if request.method == 'POST':
            selected_role = request.POST.get('selected_role')
            if selected_role in user_roles:
                request.session['selected_role'] = selected_role
                if selected_role == 'R06':
                    return redirect('faculty_dashboard')
                elif selected_role == 'R09':
                    return redirect('student_interface')
                elif selected_role in ['R01', 'R03']:
                    return redirect('admin_panel')  # Update this with the appropriate URL name for admin panel
                elif selected_role == 'R07':
                    return redirect('student_affairs_dashboard')
            else:
                messages.error(request, 'Invalid role selection.')

        return render(request, 'choose_role.html', {'user_roles': MIITRole.objects.filter(RoleID__in=user_roles)})

    elif len(user_roles) == 1 and user_roles[0].RoleID == 'R09':
        request.session['selected_role'] = 'R09'
        return redirect('student_interface')  
    
    elif len(user_roles) == 1 and user_roles[0].RoleID == 'R06':
        request.session['selected_role'] = 'R06'
        return redirect('faculty_dashboard')
    
    elif len(user_roles) == 1 and user_roles[0].RoleID == 'R07':
        request.session['selected_role'] = 'R07'
        return redirect('student_affairs_dashboard')
    
    else:
        messages.error(request, 'You do not have any roles assigned.', 'no_roles')
        return redirect('login')


def get_user_roles(email):
    try:
        # Retrieve the user object based on the email
        user = MIITUsers.objects.get(username=email)
        # Retrieve the role IDs associated with the user
        user_role_ids = MIITUserRole.objects.filter(UserID=user).values_list('RoleID', flat=True)
        # Retrieve the role descriptions based on the role IDs
        roles = MIITRole.objects.filter(RoleID__in=user_role_ids).values_list('RoleDescription', flat=True)
        return roles
    except MIITUsers.DoesNotExist:
        # Handle the case where the user does not exist
        return []

def super_admin_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        valid_roles = MIITUserRole.objects.filter(UserID=request.user.UserID).values_list('RoleID', flat=True)
        selected_role = request.session.get('selected_role')
        if selected_role in valid_roles and selected_role in ['R01','R03']:
            return view_func(request, *args, **kwargs)
        else:
            return HttpResponseForbidden("You don't have permission to access this page.")
    return _wrapped_view


def student_affairs_role_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        valid_roles = MIITUserRole.objects.filter(UserID=request.user.UserID).values_list('RoleID', flat=True)
        selected_role = request.session.get('selected_role')
        if selected_role in valid_roles and selected_role in ['R07']:
            return view_func(request, *args, **kwargs)
        else:
            return HttpResponseForbidden("You don't have permission to access this page.")
    return _wrapped_view


def student_role_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        valid_roles = MIITUserRole.objects.filter(UserID=request.user.UserID).values_list('RoleID', flat=True)
        selected_role = request.session.get('selected_role')
        if selected_role in valid_roles and selected_role in ['R09']:
            return view_func(request, *args, **kwargs)
        else:
            return HttpResponseForbidden("You don't have permission to access this page.")
    return _wrapped_view



def student_affairs_super_admin_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        valid_roles = MIITUserRole.objects.filter(UserID=request.user.UserID).values_list('RoleID', flat=True)
        selected_role = request.session.get('selected_role')
        if selected_role in valid_roles and selected_role in ['R07', 'R01']:
            return view_func(request, *args, **kwargs)
        else:
            return HttpResponseForbidden("You don't have permission to access this page.")
    return _wrapped_view


def faculty_role_required(view_func):
    def _wrapped_view(request, *args, **kwargs):
        valid_roles = MIITUserRole.objects.filter(UserID=request.user.UserID).values_list('RoleID', flat=True)
        selected_role = request.session.get('selected_role')
        if selected_role in valid_roles and selected_role in ['R06']:
            return view_func(request, *args, **kwargs)
        else:
            return HttpResponseForbidden("You don't have permission to access this page.")
    return _wrapped_view




@login_required
def change_password(request):
    if request.method == 'POST':
        current_password = request.POST['current_password']
        new_password = request.POST['new_password']
        
        user = request.user

        if not user.check_password(current_password):
            messages.error(request, 'Current password is incorrect.', extra_tags='incorrect')
            return redirect('change_password')

        user.set_password(new_password)
        user.save()

        update_session_auth_hash(request, user)  # Important to keep the user logged in after password change
        messages.success(request, 'Your password has been successfully changed.', extra_tags='success')
        return redirect('change_password')

    return render(request, 'change_password.html')




def validate_captcha(captcha_response):
    data = {
        'secret': settings.RECAPTCHA_PRIVATE_KEY,
        'response': captcha_response
    }
    response = requests.post('https://www.google.com/recaptcha/api/siteverify', data=data)
    result = response.json()
    return result.get('success')

def forgot_password(request):
    if request.method == 'POST':
        email = request.POST['email']
        captcha_response = request.POST.get('g-recaptcha-response')
        
        if captcha_response and validate_captcha(captcha_response) and MIITUsers.objects.filter(username=email).exists():
            token_entry = PasswordResetToken.objects.filter(email=email).first()
            if token_entry:
                if token_entry.expires_at > timezone.now():
                    messages.info(request, 'A password reset link has already been sent to your email. Please check your inbox.', extra_tags='info')
                    return redirect('forgot_password')
                else:
                    token_entry.delete()

            token = get_random_string(50)
            token_hash = bcrypt.hashpw(token.encode(), bcrypt.gensalt()).decode()
            PasswordResetToken.objects.create(email=email, token_hash=token_hash)

            reset_link = request.build_absolute_uri(
                reverse('reset_password') + f"?token={urlsafe_base64_encode(force_bytes(token))}&email={email}"
            )
            send_mail(
                'Password Reset Request',
                f'''Dear {email},

We received a request to reset your password. Please click the link below to reset your password:

{reset_link}

If you did not request this password reset, please ignore this email. This link will expire in 15 minutes.

Thank you,
MIIT Academia''',
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
            )
            messages.success(request, f'Password reset link has been sent to {email}.', extra_tags='success')
            return redirect('forgot_password')
        else:
            messages.error(request, 'Invalid email address or CAPTCHA.', extra_tags='error')

    context = {'RECAPTCHA_PUBLIC_KEY': settings.RECAPTCHA_PUBLIC_KEY}
    return render(request, 'forgot_password.html', context)


def reset_password(request):
    if request.method == 'POST':
        token = force_str(urlsafe_base64_decode(request.GET.get('token')))
        email = request.GET.get('email')
        new_password = request.POST['new_password']
        token_entry = PasswordResetToken.objects.filter(email=email).first()
        if token_entry and bcrypt.checkpw(token.encode(), token_entry.token_hash.encode()):
            if token_entry.expires_at > timezone.now():
                user = MIITUsers.objects.get(username=email)
                user.set_password(new_password)
                user.save()
                token_entry.delete()
                messages.success(request, 'Password has been reset successfully.', extra_tags='reset')
                return redirect('login')
            else:
                messages.error(request, 'The password reset link has expired.', extra_tags='error')
        else:
            messages.error(request, 'Invalid or expired token.', extra_tags='error')

    return render(request, 'reset_password.html')


@receiver(user_logged_in)
def log_user_login(sender, request, user, **kwargs):
    ip_address = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', '')

    LoginHistory.objects.create(
        user=user,
        ip_address=ip_address,
        user_agent=user_agent
    )

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip



@login_required
def login_history(request):
    history = LoginHistory.objects.filter(user=request.user).order_by('-login_time')
    return render(request, 'login_history.html', {'history': history})


@login_required
def home(request):
    # If the user has selected a role, redirect accordingly
    selected_role = request.session.get('selected_role')
    if selected_role:
        if selected_role == 'R06':
            return redirect('faculty_dashboard')
        elif selected_role == 'R09':
            return redirect('student_interface')
        elif selected_role in ['R01', 'R03']:
            return redirect('admin_panel')  # Update with the appropriate URL name for admin panel
        elif selected_role == 'R07':
            return redirect('student_affairs_dashboard')

    # If no role is selected, handle role selection
    user_roles = request.session.get('user_roles', [])
    if len(user_roles) > 1:
        return redirect('choose_role')
    elif len(user_roles) == 1:
        return redirect_by_role(user_roles[0])
    else:
        messages.error(request, 'You do not have any roles assigned.')
        return render(request, 'login.html')
    
def redirect_by_role(role_id,request):
    if role_id == 'R09':
        return redirect('student_interface')
    elif role_id == 'R06':
        return redirect('faculty_dashboard')
    elif role_id in ['R01', 'R03']:
        return redirect('admin_panel')  # Update with the appropriate URL name for admin panel
    elif role_id == 'R07':
        return redirect('student_affairs_dashboard')
    else:
        # Handle other roles
        return render(request, 'login.html')  # Render login page or handle accordingly


@login_required
def profile(request):
    selected_user = None
    selected_user_type = None
    program = None
    discipline = None
    batch = None

    try:
        selected_user = Student.objects.get(UserID=request.user.UserID)
        selected_user_type = 'Student'
        program=  Program.objects.filter(ProgramID=selected_user.ProgramID).values_list('ProgramFullName', flat=True).first(),
        discipline = Discipline.objects.filter(DisciplineID=selected_user.DisciplineID).values_list('DisciplineFullName', flat=True).first(),
        batch = Batch.objects.filter(BatchID=selected_user.BatchID).values_list('BatchYear', flat=True).first(),

    except Student.DoesNotExist:
        try:
            selected_user = FacultyStaff.objects.get(UserID=request.user.UserID)
            selected_user_type = 'Faculty'
        except FacultyStaff.DoesNotExist:
            messages.error(request, 'No user found with this email.', extra_tags='not_found')

    context = {
        'selected_user': selected_user,
        'selected_user_type': selected_user_type,
        'program':program,
        'discipline':discipline,
        'batch':batch,
    }
    return render(request, 'profile.html', context)


@login_required
@super_admin_required
def admin_panel(request):
    # print(request.session['selected_role'])
    return render(request, 'admin_panel.html')

@login_required
@super_admin_required
def check_database(request):
    tables_data = {}
    with connection.cursor() as cursor:
        cursor.execute("SHOW TABLES")
        tables = cursor.fetchall()
        for table in tables:
            table_name = table[0]
            cursor.execute(f"SELECT * FROM {table_name}")
            table_data = cursor.fetchall()
            tables_data[table_name] = table_data
    
    return render(request, 'check_database.html', {'tables_data': tables_data})


@login_required
@super_admin_required
def table_data(request, table_name):
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM {table_name}")
    table_data = cursor.fetchall()
    column_names = [desc[0] for desc in cursor.description]
    return render(request, 'table_data.html', {'table_name': table_name, 'column_names': column_names, 'table_data': table_data})


@login_required
@super_admin_required
def user_management(request):
    return render(request, 'user_management.html')



@login_required
@super_admin_required
def create_student(request):
    form = None
    if request.method == 'POST':
        if 'create_one_by_one' in request.POST:
            form = StudentForm(request.POST)
            if form.is_valid():
                # Extract email from form.cleaned_data
                email = form.cleaned_data.get('email')

                # Create a new user in the MIITUsers table
                user_id = generate_user_id()

                try:
                    user = MIITUsers.objects.create(
                    UserID=user_id,
                    username=email,
                    UserPasswordKey=hash_password('Thanks123!'),
                    UserStatus='Active'
                    )
                
                except Exception as e:
                    messages.error(request, e, extra_tags='error')
                    return redirect('create_student')


                # Create the student entry
                student = form.save(commit=False)
                student.UserID = user
                student.ACBStatus = False
                student.StudentID = generate_student_id()  # Implement this function to generate StudentID
                student.save()

                role = MIITRole.objects.get(RoleID='R09')

                MIITUserRole.objects.create(
                    UserID=user,
                    RoleID=role
                )

                messages.success(request, 'Student account created successfully.')
                return redirect('create_student')
            
        elif 'bulk_registration' in request.POST:
            bulk_form = BulkRegistrationForm(request.POST, request.FILES)
            if bulk_form.is_valid():
                excel_file = request.FILES.get('excel_file')
                if excel_file:
                    try:
                        # Read data from the Excel file
                        df = pd.read_excel(excel_file)

                        # Iterate through each row and create student accounts
                        for index, row in df.iterrows():
                            email = row['Email']
                            # Create a new user in the MIITUsers table
                            user_id = generate_user_id()
                            user = MIITUsers.objects.create(
                                UserID=user_id,
                                username=email,
                                UserPasswordKey=hash_password('Thanks123!'),
                                UserStatus='Active'
                            )
                            student = Student(
                                StudentID=generate_student_id(),  # Implement this function to generate StudentID
                                UserID=user,
                                DisciplineID=Discipline.objects.filter(DisciplineShortName=row['Discipline']).first(),
                                ProgramID=Program.objects.filter(ProgramShortName=row['Program']).first(),
                                BatchID=Batch.objects.filter(BatchYear=row['Batch Year']).first(),
                                StudentName=row['Student Name'],
                                Salutation=row['Salutation'],
                                SectionName=row['Section'],
                                RollNumber=row['Roll Number'],
                                StudentNRC=row['Student NRC'],
                                StudentPhone=row['Student Phone'],
                                StudentDOB=row['Student DOB'],
                                Nationality=row['Nationality'],
                                Religion=row['Religion'],
                                MatricRollNumber=row['MatricRollNumber'],
                                MatricExamYear=row['Matric Exam Year'],
                                ACBStatus=False,
                                FatherName=row['FatherName'],
                                FatherNRC=row['FatherNRC'],
                                FatherPhoneNumber=row['FatherPhoneNumber'],
                                MotherName=row['MotherName'],
                                MotherNRC=row['MotherNRC'],
                                MotherPhoneNumber=row['MotherPhoneNumber'],
                                Address=row['Address']
                            )
                            student.save()

                            role = MIITRole.objects.get(RoleID='R09')

                            MIITUserRole.objects.create(
                                UserID=user,
                                RoleID=role
                            )

                        messages.success(request, 'Students registered successfully.')
                        return redirect('create_student')
                    except Exception as e:
                        messages.error(request, f"Error processing the Excel file: {e}")

    else:
        form = StudentForm()
        bulk_form = BulkRegistrationForm()
    
    disciplines = Discipline.objects.all()
    programs = Program.objects.all()
    batches = Batch.objects.all()

    context = {
        'form': form,
        'bulk_form': bulk_form,
        'disciplines': disciplines,
        'programs': programs,
        'batches': batches
    }
    return render(request, 'create_student.html', context)

@login_required
@super_admin_required
def create_faculty(request):
    if request.method == 'POST':
        if 'create_one_by_one' in request.POST:
            form = FacultyStaffForm(request.POST)
            if form.is_valid():
                email = form.cleaned_data.get('email')
                user_id = generate_user_id()

                try:
                    user = MIITUsers.objects.create(
                        UserID=user_id,
                        username=email,
                        UserPasswordKey=hash_password('Thanks123!'),
                        UserStatus='Active'
                    )
                except Exception as e:
                    messages.error(request, e, extra_tags='error')
                    return redirect('create_faculty')

                faculty = form.save(commit=False)
                faculty.UserID = user
                faculty.FacultyStaffID = generate_faculty_id()  # Implement this function to generate FacultyStaffID
                faculty.save()

                role = MIITRole.objects.get(RoleID='R06')  # Assuming 'R02' is the role ID for faculty staff

                MIITUserRole.objects.create(
                    UserID=user,
                    RoleID=role
                )

                messages.success(request, 'Faculty account created successfully.')
                return redirect('create_faculty')

        elif 'bulk_registration' in request.POST:
            bulk_form = BulkRegistrationForm(request.POST, request.FILES)
            if bulk_form.is_valid():
                excel_file = request.FILES.get('excel_file')
                if excel_file:
                    try:
                        df = pd.read_excel(excel_file)
                        for index, row in df.iterrows():
                            email = row['Email']
                            user_id = generate_user_id()
                            user = MIITUsers.objects.create(
                                UserID=user_id,
                                username=email,
                                UserPasswordKey=hash_password('Thanks123!'),
                                UserStatus='Active'
                            )
                            faculty = FacultyStaff(
                                FacultyStaffID=generate_faculty_id(),  # Implement this function
                                FacultyStaffName=row['Faculty Staff Name'],
                                ShortName=row['Short Name'],
                                Salutation=row['Salutation'],
                                Designation=row['Designation'],
                                Department=row['Department'],
                                NRC=row['NRC'],
                                Phone=row['Phone'],
                                Address=row['Address'],
                                UserID=user
                            )
                            faculty.save()

                            role = MIITRole.objects.get(RoleID='R02')
                            MIITUserRole.objects.create(
                                UserID=user,
                                RoleID=role
                            )
                        messages.success(request, 'Bulk faculty registration successful.')
                        return redirect('create_faculty')
                    except Exception as e:
                        messages.error(request, f'Error processing Excel file: {e}', extra_tags='error')
                        return redirect('create_faculty')
            else:
                messages.error(request, 'Invalid file format.', extra_tags='error')
                return redirect('create_faculty')
    else:
        form = FacultyStaffForm()
        bulk_form = BulkRegistrationForm()

    context = {
        'form': form,
        'bulk_form': bulk_form,
    }

    return render(request, 'create_faculty.html', context)


def generate_user_id():
    # Logic to generate the next UserID
    last_user = MIITUsers.objects.order_by('UserID').last()
    if not last_user:
        return 'U00001'
    user_id = int(last_user.UserID[1:]) + 1
    return f'U{user_id:05d}'

def generate_faculty_id():
    # Logic to generate the next StudentID
    last_faculty = FacultyStaff.objects.order_by('FacultyStaffID').last()
    if not last_faculty:
        return 'F0001'
    faculty_id = int(last_faculty.FacultyStaffID[1:]) + 1
    return f'F{faculty_id:04d}'

def generate_student_id():
    # Logic to generate the next StudentID
    last_student = Student.objects.order_by('StudentID').last()
    if not last_student:
        return 'STU00001'
    student_id = int(last_student.StudentID[3:]) + 1
    return f'STU{student_id:05d}'

def hash_password(password):
    ph = PasswordHasher(time_cost=4, memory_cost=65536)
    return ph.hash(password)


@login_required
@super_admin_required
def academic_system_management(request):
    return render(request, 'academic_system_management.html')

@login_required
@super_admin_required
def academic_settings(request):
    return render(request, 'academic_settings.html')

@login_required
@super_admin_required
def create_academic_year(request):
    if request.method == 'POST':
        # Get form data
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        # Validate form data
        if not start_date or not end_date:
            messages.error(request, 'Please select both start date and end date.')
            return redirect('create_academic_year')

        try:
            # Calculate academic year ID
            existing_count = AcademicYear.objects.count()
            new_ay_id = f'AY{existing_count + 1:02}'

            # Create AcademicYear object
            academic_year = AcademicYear(
                AY_ID=new_ay_id,
                AYStartDate=start_date,
                AYEndDate=end_date,
                AYCreateDate=date.today()
            )
            academic_year.save()

            # Display success message
            messages.success(request, 'Academic year created successfully.')
            return redirect('create_academic_year')  # Redirect to the same page after submission
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
            return redirect('create_academic_year')
    else:
        return render(request, 'create_academic_year.html')

@login_required
@super_admin_required
def edit_academic_year(request):
    academic_years = AcademicYear.objects.all()
    selected_academic_year_id = request.GET.get('academic_year_id')
    
    if selected_academic_year_id:
        print(selected_academic_year_id)
        academic_year = AcademicYear.objects.get(AY_ID=selected_academic_year_id)
        form = AYForm(request.POST or None, instance=academic_year)
    else:
        form = None
    
    if form and form.is_valid():
        form.save()
        messages.success(request, "Academic Year updated successfully.")
    
    context = {
        'academic_years': academic_years,
        'selected_academic_year_id': selected_academic_year_id,
        'form': form,
    }
    return render(request, 'edit_academic_year.html', context)    

@login_required
@super_admin_required
def create_program(request):
    program_count = Program.objects.count()+1
    form = ProgramForm(request.POST)
    if form.is_valid():
        program = form.save(commit = False)
        program.ProgramID = f"P{program_count:02d}"
        program.save()
        messages.success(request, "New Program Create successfully.")
        return redirect('create_program')
    else:
        form = ProgramForm()

    context={
        'form' : form,
    }
    return render(request, 'create_program.html', context)

@login_required
@super_admin_required
def edit_program(request):
    programs = Program.objects.all()
    program_id = request.GET.get('program_id')
    
    if program_id:
        program = Program.objects.get(ProgramID=program_id)
        form = ProgramForm(request.POST or None, instance=program)
    else:
        form = None
    
    if form and form.is_valid():
        form.save()
        messages.success(request, "Program updated successfully.")
    
    context = {
        'form': form,
        'programs': programs,
        'selected_program_id': program_id
    }
    return render(request, 'edit_program.html', context)

@login_required
@super_admin_required
def create_discipline(request):
    discipline_count = Discipline.objects.count() + 1
    form = DisciplineForm(request.POST)
    if form.is_valid():
        discipline = form.save(commit=False)
        discipline.DisciplineID = f"D{discipline_count:02d}"
        discipline.save()
        messages.success(request, 'Create new Discipline successfully.')
        return redirect('create_discipline')
    else:
        form = DisciplineForm()
   
    context={
        'form': form,
    }
    return render(request, 'create_discipline.html', context)
    
@login_required
@super_admin_required
def edit_discipline(request):
    disciplines = Discipline.objects.all()
    discipline_id = request.GET.get('discipline_id')
    
    if discipline_id:
        discipline = Discipline.objects.get(DisciplineID=discipline_id)
        form = DisciplineForm(request.POST or None, instance=discipline)
    else:
        form = None
    
    if form and form.is_valid():
        form.save()
        messages.success(request, "Discipline updated successfully.")
    
    context = {
        'form': form,
        'disciplines': disciplines,
        'selected_discipline_id': discipline_id
    }
    return render(request, 'edit_discipline.html', context)

@login_required
@super_admin_required
def create_semester(request):
    ay = AcademicYear.objects.filter(AYStatus='Current')
    programs = Program.objects.all()
    if request.method == 'POST':
        program_id = request.POST.get("program")
        ay_id = request.POST.get('academic_year')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        description = request.POST.get('description')
        academic_year = AcademicYear.objects.get(AY_ID=ay_id)
        program = Program.objects.get(ProgramID=program_id)

        # Validate form data
        if not start_date or not end_date:
            messages.info(request, 'Please select both start date and end date.',extra_tags='notice')
            return redirect('create_semester')

        try:
            existing_count = Semester.objects.count()
            new_semester_id = f'SEM{existing_count + 1:03}'
            # Create Semester object
            semester = Semester(
                SemesterID=new_semester_id,
                AY_ID=academic_year,
                ProgramID=program,
                SemesterStatus='New',
                SemesterStartDate=start_date,
                SemesterEndDate=end_date,
                SemesterDescription = description,
            )
            semester.save()
            
            # Display success message
            messages.success(request, 'New Semester created successfully.',extra_tags='success')
            return redirect('create_semester')  # Redirect to the same page after submission
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}',extra_tags='error')
            return redirect('create_semester')
    else:
        context = {
            'ay': ay,
            'programs': programs,
        }
        return render(request, 'create_semester.html', context)

@login_required
@super_admin_required
def edit_semester(request):
    academic_years = AcademicYear.objects.all()
    selected_academic_year_id = request.GET.get('academic_year')
    selected_semester_id = request.GET.get('semester')
 
    semesters = Semester.objects.filter(AY_ID=selected_academic_year_id) if selected_academic_year_id else []
    semester = get_object_or_404(Semester, SemesterID=selected_semester_id) if selected_semester_id else None

    if request.method == 'POST' and semester:
        form = SemesterForm(request.POST, instance=semester)
        if form.is_valid():
            form.save()
            messages.success(request, 'Semester details updated successfully.')
            return redirect(f'{request.path}?academic_year={selected_academic_year_id}&semester={selected_semester_id}')
    else:
        form = SemesterForm(instance=semester) if semester else None

    context = {
        'academic_years': academic_years,
        'semesters': semesters,
        'form': form,
        'selected_academic_year_id': selected_academic_year_id,
        'selected_semester_id': selected_semester_id
    }
    return render(request, 'edit_semester.html', context)


@login_required
@super_admin_required
def get_semesters(request):
    academic_year_id = request.GET.get('academic_year')
    semesters = Semester.objects.filter(AY_ID=academic_year_id).values('SemesterID', 'SemesterDescription')
    return JsonResponse(list(semesters), safe=False)



@login_required
@super_admin_required
def academic_statistics(request):
    return render(request, 'academic_statistics.html')



@login_required
@super_admin_required
def view_base_chart(request):
    if request.method == 'POST':
        selected_batch_id = request.POST.get('selected_batch')
        selected_discipline_id = request.POST.get('selected_discipline')

        # Fetch base chart data for the selected batch and discipline
        base_chart_data = get_base_chart_data(selected_batch_id, selected_discipline_id)
        # Fetch all available disciplines for the selected batch
        number = ['','I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X']
        context = {
            'batches': Batch.objects.all(),
            'selected_batch': selected_batch_id,
            'selected_discipline': selected_discipline_id,
            'base_chart_data': base_chart_data,
            'disciplines': Discipline.objects.all(),
            'number':number,
        }
        return render(request, 'view_basechart.html', context)
    else:
        # Render the initial page with batch and discipline selection form
        context = {
            'batches': Batch.objects.all(),
            'selected_batch': None,
            'selected_discipline': None,
            'base_chart_data': None,
            'disciplines': Discipline.objects.all(),
        }
        return render(request, 'view_basechart.html', context)


def get_base_chart_data(batch_id, discipline_id):
    years = OfferedCourses.objects.filter(
        BaseChartID__BatchID=batch_id,
        BaseChartID__DisciplineID=discipline_id
    ).values_list('YearNumber', flat=True).distinct().order_by('YearNumber')
    
    base_chart_data = {}
    
    for year in years:
        semester_course = {}
        semester_numbers = list(OfferedCourses.objects.filter(
            BaseChartID__BatchID=batch_id,
            BaseChartID__DisciplineID=discipline_id,
            YearNumber=year
        ).values_list('SemesterPeriodNumber', flat=True).distinct().order_by('SemesterPeriodNumber'))
        
        semester_numbers = sorted(semester_numbers, key=lambda x: int(x[1:]))

        for semester_number in semester_numbers:
            courses = Course.objects.filter(
                offeredcourses__SemesterPeriodNumber=semester_number,
                offeredcourses__YearNumber=year,
                offeredcourses__BaseChartID__BatchID=batch_id,
                offeredcourses__BaseChartID__DisciplineID=discipline_id
            ).values('CourseNumber', 'CourseName', 'CourseCredits')
            semester_course[semester_number] = list(courses)
        
        base_chart_data[year] = semester_course
    
    return base_chart_data


def get_disciplines(batch_id):
    disciplines = ProgramBatchDiscipline.objects.filter(BatchID=batch_id).values('DisciplineID').distinct()
    return disciplines



@login_required
@super_admin_required
def broadcast_message(request):
    if request.method == 'POST':
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        recipients = request.POST.getlist('recipients')

        email_messages = []
        recipient_count = {
            'students': 0,
            'faculty_staff': 0,
            'all': 0,
        }

        if 'students' in recipients or 'all' in recipients:
            for student in Student.objects.all():
                try:
                    user = MIITUsers.objects.get(username=student.UserID)
                except:
                    continue
                
                if user.UserStatus == 'active':
                    print(f'Email : {user.username}')
                    email_messages.append(
                        (subject, message, settings.DEFAULT_FROM_EMAIL, ['hlaminnaing013@gmail.com'])
                    )
                    email_messages.append(
                        (subject, message, settings.DEFAULT_FROM_EMAIL, ['acseguywithlife404@gmail.com'])
                    )
                    break

            recipient_count['students'] = len(email_messages)

        if 'faculty_staff' in recipients or 'all' in recipients:
            for faculty_staff in FacultyStaff.objects.all():
                try:
                    user = MIITUsers.objects.get(username=faculty_staff.UserID)
                except:
                    continue
                
                email_messages.append(
                    (subject, message, settings.DEFAULT_FROM_EMAIL, [user.username])
                )
            recipient_count['faculty_staff'] = len(email_messages) - recipient_count['students']

        recipient_count['all'] = len(email_messages)

        try:
            send_mass_mail(tuple(email_messages), fail_silently=False)
            messages.success(request, 'Message sent successfully.')
        except Exception as e:
            messages.error(request, f'Error sending message: {str(e)}')

        context = {
            'recipient_count': recipient_count,
        }
        return render(request, 'broadcast_message_sent.html', context)

    return render(request, 'broadcast_message.html')


@login_required
@faculty_role_required
def faculty_dashboard(request):
    # Dummy data for courses previously taught and currently being taught
    previous_courses = [
        "Introduction to Programming",
        "Data Structures and Algorithms",
        "Computer Networks",
        "Database Management Systems",
        "Operating Systems",
        "Software Engineering"
    ]
    current_courses = [
        "Machine Learning",
        "Web Development",
        "Artificial Intelligence",
        "Cybersecurity"
    ]

    context = {
        'previous_courses': previous_courses,
        'current_courses': current_courses
    }
    return render(request, 'faculty_dashboard.html', context)


@login_required
@faculty_role_required
def faculty_courses(request):
    ay_courses = defaultdict(lambda: defaultdict(list))
    
    # Get the current teacher's ID
    user_id = request.user.UserID
    facultystaff_id = FacultyStaff.objects.filter(UserID=user_id).values_list('FacultyStaffID', flat=True).first()
    
    # Get the semester ID
    current_semester = Semester.objects.filter(SemesterStatus='Current').values_list('SemesterID', flat=True).first()
    previous_semesters = Semester.objects.filter(SemesterStatus='Completed').values_list('SemesterID', flat=True)
    academic_years = Semester.objects.filter(SemesterID__in=previous_semesters).values_list('AY_ID', 'SemesterID')
    
    # Get the courses taught by the current teacher in the current semester
    current_courses = FacultyCourse.objects.filter(FacultyStaffID=facultystaff_id, SemesterID=current_semester, Deleted=False).values('CourseNumber', 'Section', 'BatchID')
    
    previous_courses = FacultyCourse.objects.filter(FacultyStaffID=facultystaff_id, SemesterID__in=previous_semesters, Deleted=False).values('CourseNumber', 'Section', 'BatchID', 'SemesterID')
    
    for course in current_courses:
        course_name = Course.objects.filter(CourseNumber=course['CourseNumber']).values_list('CourseName', flat=True).first()
        course['CourseName'] = course_name
    
    for course in previous_courses:
        course_name = Course.objects.filter(CourseNumber=course['CourseNumber']).values_list('CourseName', flat=True).first()
        semester_id = course['SemesterID']
        ay_id = [ay for ay, sem in academic_years if sem == semester_id][0]
        ay_courses[ay_id][semester_id].append({
            'CourseNumber': course['CourseNumber'],
            'CourseName': course_name,
            'Section': course['Section'],
            'BatchID': course['BatchID'],
        })

        # Fetch AY start and end dates
        start_date = AcademicYear.objects.filter(AY_ID=ay_id).values_list('AYStartDate', flat=True).first().year
        end_date = AcademicYear.objects.filter(AY_ID=ay_id).values_list('AYEndDate', flat=True).first().year

        # Include AY start and end dates in the dictionary
        ay_courses[ay_id]['AYStartDate'] = start_date
        ay_courses[ay_id]['AYEndDate'] = end_date
    
    # Convert inner defaultdicts to regular dicts
    ay_courses = {k: dict(v) for k, v in ay_courses.items()}
   
    context = {
        'current_courses': list(current_courses),
        'previous_courses': ay_courses,
    }
    
    return render(request, 'faculty_courses.html', context)
    



@login_required
@faculty_role_required
def current_courses(request):
    # Get the current teacher's ID
    user_id = request.user.UserID
    facultystaff_id = FacultyStaff.objects.filter(UserID=user_id).values_list('FacultyStaffID', flat=True).first()
    
    # Get the current semester ID
    current_semester = Semester.objects.filter(SemesterStatus='Current').values_list('SemesterID', flat=True).first()
    
    # Get the courses taught by the current teacher in the current semester
    courses = FacultyCourse.objects.filter(FacultyStaffID=facultystaff_id, SemesterID=current_semester, Deleted=False).values('CourseNumber', 'Section', 'BatchID')
    for course in courses:
        course_name = Course.objects.filter(CourseNumber=course['CourseNumber']).values_list('CourseName', flat=True).first()
        course['CourseName'] = course_name

    context = {
        'courses': courses,
    }
    return render(request, 'current_courses.html', context)

@login_required
@faculty_role_required
def faculty_projects(request):
    faculty_id = FacultyStaff.objects.filter(UserID=request.user.UserID).values_list('FacultyStaffID', flat=True).first()
    current_semesters = Semester.objects.filter(SemesterStatus='Current')
    current_projects = Project.objects.filter(FacultyStaffID=faculty_id, SemesterID__in=current_semesters, Deleted=False).order_by('ProgramID')
    previous_projects = Project.objects.filter(FacultyStaffID=faculty_id).exclude(SemesterID__in=current_semesters).filter(Deleted=False).order_by('ProgramID')

    if request.method == 'POST':
        form = ProjectProposalForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.FacultyStaffID = FacultyStaff.objects.get(FacultyStaffID=faculty_id)
            project.SemesterID = current_semesters.first()
            project.ProjectNumber = generate_project_number()
            project.save()
            messages.success(request, 'Project proposed successfully.')
            return redirect('faculty_projects')
    else:
        form = ProjectProposalForm()  # Allow faculty to enter the credits value
    context = {
        'current_projects': group_projects_by_program(current_projects),
        'previous_projects': group_projects_by_semester_and_program(previous_projects),
        'form': form
    }
    print(group_projects_by_semester_and_program(previous_projects))
    return render(request, 'faculty_projects.html', context)

def generate_project_number():
    last_project = Project.objects.order_by('-ProjectNumber').first()
    if not last_project:
        return 'P0001'
    last_id = int(last_project.ProjectNumber[1:])
    new_id = f'P{last_id + 1:04d}'
    return new_id

def group_projects_by_program(projects):
    grouped = {}
    for project in projects:
        program_name = project.ProgramID.ProgramShortName
        if program_name not in grouped:
            grouped[program_name] = []
        grouped[program_name].append(project)
    return grouped

def group_projects_by_semester_and_program(projects):
    grouped = {}
    for project in projects:
        semester_id = project.SemesterID
        program_name = project.ProgramID.ProgramShortName

        if program_name not in grouped:
            grouped[program_name] = {}
        
        if semester_id not in grouped[program_name]:
            grouped[program_name][semester_id] = []
        
        grouped[program_name][semester_id].append(project)
    return grouped


@login_required
@faculty_role_required
def upload_attendance(request, course_code):
    faculty = FacultyStaff.objects.get(UserID=request.user.UserID)
    semester = Semester.objects.get(SemesterStatus='Current')
    course = Course.objects.get(CourseNumber=course_code)

    attendance = Attendance.objects.filter(
        FacultyStaffID=faculty,
        SemesterID=semester,
        CourseNumber=course
    ).order_by('-EndDate').first()  # Get the latest attendance record

    start_date = attendance.StartDate if attendance else None
    end_date = attendance.EndDate if attendance else None
    end_date_temp = end_date

    if request.method == 'POST':
        if 'attendance_file' in request.FILES:
            attendance_file = request.FILES['attendance_file']
            if attendance_file.name.endswith('.xlsx'):
                # Process the uploaded file
                try:
                    workbook = load_workbook(attendance_file)
                    sheet = workbook['Attendance_Input']

                    # Extract data
                    start_date = sheet.cell(row=4, column=5).value.split('All')[0].strip()
                    end_date = sheet.cell(row=4, column=sheet.max_column - 6).value.split('All')[0].strip()

                    # Ensure the input format matches exactly
                    try:
                        start_date_obj = datetime.datetime.strptime(start_date, "%d %b %Y %I.%M%p").date()
                        end_date_obj = datetime.datetime.strptime(end_date, "%d %b %Y %I.%M%p").date()
                    except ValueError as e:
                        messages.error(request, f"Error processing date: {e}", extra_tags='error')
                        return redirect('upload_attendance', course_code)

                    # Get Attendance ID or create a new one
                    with transaction.atomic():
                        attendance, created = Attendance.objects.get_or_create(
                            FacultyStaffID=faculty,
                            SemesterID=semester,
                            CourseNumber=course,
                            defaults={'StartDate': start_date_obj, 'EndDate': end_date_obj}
                        )


                        if attendance.EndDate == end_date_obj and attendance.StartDate == start_date_obj and not created:
                            messages.add_message(request, messages.INFO, f'Attendance already uploaded till this date : {end_date_temp if end_date_temp else ""}.', extra_tags='notice')
                            return redirect('upload_attendance', course_code)

                        if attendance.EndDate > start_date_obj and not created:
                            messages.add_message(request, messages.INFO, f'Attendance date might be overlapping. Please upload after this date : {end_date_temp if end_date_temp else ""}.', extra_tags='notice')
                            return redirect('upload_attendance', course_code)

                        # Update start and end dates if needed
                        if not created:
                            attendance.StartDate = min(attendance.StartDate, start_date_obj)
                            attendance.EndDate = max(attendance.EndDate, end_date_obj)
                            attendance.save()

                        for row in sheet.iter_rows(min_row=5, values_only=True):
                            student_id = row[1]
                            total_sections = row[-3]
                            present_sections = row[-6]

                            user_id = MIITUsers.objects.filter(username=student_id).values_list('UserID', flat=True).first()
                            student = Student.objects.get(UserID=user_id)

                            student_attendance, sa_created = StudentAttendance.objects.get_or_create(
                                AttendanceID=attendance,
                                StudentID=student,
                                BatchID=student.BatchID,
                                defaults={
                                    'TotalSections': total_sections,
                                    'PresentSections': present_sections,
                                    'Percentage': (present_sections / total_sections) * 100
                                }
                            )

                            if not sa_created:
                                student_attendance.TotalSections += total_sections
                                student_attendance.PresentSections += present_sections
                                student_attendance.Percentage = (student_attendance.PresentSections / student_attendance.TotalSections) * 100
                                student_attendance.save()

                    messages.success(request, 'Attendance file processed and uploaded successfully.', extra_tags='success')
                except Exception as e:
                    messages.error(request, f'Error processing file: {e}', extra_tags='error')
                return redirect('upload_attendance', course_code)
            else:
                messages.error(request, 'Please upload a valid .xlsx file.', extra_tags='error')
        else:
            messages.error(request, 'No file uploaded. Please upload an attendance file.', extra_tags='error')

    context = {
        'course_id': course_code,
        'courseName' : Course.objects.filter(CourseNumber=course_code).values_list('CourseName', flat=True).first(),
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'upload_attendance.html', context)

@login_required
def settings_view(request):
    mfa_enabled = MIITUsers.objects.filter(UserID=request.user.UserID).values_list('mfa_enabled', flat=True).first()
    mfa_status = 'Enabled' if mfa_enabled else 'Disabled'
    mfa_button_text = 'Disable MFA' if mfa_enabled else 'Enable MFA'
    
    context = {
        'mfa_status': mfa_status,
        'mfa_button_text': mfa_button_text
    }
    
    return render(request, 'settings.html', context)

@login_required
def initialize_mfa_setup(request):
    user = request.user
    user_profile = MIITUsers.objects.get(UserID=user.UserID)
    user_profile.generate_mfa_secret()
    user_profile.save()
    totp = pyotp.TOTP(user_profile.mfa_secret)
    otp_auth_url = totp.provisioning_uri(user.username, issuer_name="MIIT Academia")
    qr_code = generate_qr_with_logo_and_frame(otp_auth_url, f"{settings.STATICFILES_DIRS[0]}/images/miit-logo.png")
    
    return JsonResponse({
        'qr_code': f'data:image/png;base64,{qr_code}',
        'secret_token': user_profile.mfa_secret
    })

@login_required
def verify_mfa_code(request):
    user = request.user
    user_profile = MIITUsers.objects.get(UserID=user.UserID)
    totp = pyotp.TOTP(user_profile.mfa_secret)
    code = json.loads(request.body).get('code')
    if totp.verify(code):
        backup_codes = user_profile.generate_backup_codes()
        user_profile.save()
        return JsonResponse({'success': True, 'backup_codes': backup_codes})
    else:
        return JsonResponse({'success': False})

@login_required
def finish_mfa_setup(request):
    user = request.user
    user_profile = MIITUsers.objects.get(UserID=user.UserID)
    user_profile.mfa_enabled = True
    user_profile.save()
    return JsonResponse({'success': True})

@login_required
def toggle_mfa(request):
    user = request.user
    user_profile = MIITUsers.objects.get(UserID=user.UserID)
    if user_profile.mfa_enabled:
        # Disable MFA
        totp = pyotp.TOTP(user_profile.mfa_secret)
        code = json.loads(request.body).get('code')
        if totp.verify(code):
            user_profile.mfa_enabled = False
            user_profile.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False})
    else:
        # Enable MFA via the multi-step process
        pass  # The actual enabling process is handled in the steps above

    return redirect('settings')


def generate_qr_with_logo_and_frame(data, logo_path):
    # Generate QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert('RGBA')

    # Create a white circle frame
    circle_radius = 100
    frame_width = 10
    frame_color = "white"
    circle_center = ((qr_img.size[0] - 1) // 2, (qr_img.size[1] - 1) // 2)
    frame_box = (
        circle_center[0] - circle_radius - frame_width,
        circle_center[1] - circle_radius - frame_width,
        circle_center[0] + circle_radius + frame_width,
        circle_center[1] + circle_radius + frame_width,
    )
    frame_img = Image.new("RGBA", qr_img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(frame_img)
    draw.ellipse(frame_box, fill=frame_color)

    # Paste frame on QR code
    qr_img = Image.alpha_composite(qr_img, frame_img)

    # Load logo
    logo = Image.open(logo_path)
    logo = logo.resize((circle_radius * 2, circle_radius * 2), Image.LANCZOS)

    # Calculate position for the logo
    position = (
        circle_center[0] - circle_radius,
        circle_center[1] - (circle_radius - 4),
    )

    # Paste logo on QR code
    qr_img.paste(logo, position, logo)

    # Convert the image to a byte stream and then to base64
    byte_arr = BytesIO()
    qr_img.save(byte_arr, format='PNG')
    byte_arr = byte_arr.getvalue()
    return base64.b64encode(byte_arr).decode('utf-8')

@login_required
def notification_settings(request):
    user = request.user
    try:
        preferences = NotificationPreference.objects.get(user=user)
    except NotificationPreference.DoesNotExist:
        preferences = NotificationPreference(user=user)
        preferences.save()

    if request.method == 'POST':
        form = NotificationPreferencesForm(request.POST, instance=preferences)
        if form.is_valid():
            form.save()
            messages.success(request, 'Notification preferences updated successfully.')
            return redirect('notification_settings')
    else:
        form = NotificationPreferencesForm(instance=preferences)

    return render(request, 'notification_settings.html', {'form': form})

@login_required
def student_interface(request):
    return render(request, 'student_interface.html')



@login_required
@student_role_required
def check_grades(request):
    student_name = None
    student_id = Student.objects.filter(UserID=request.user.UserID).values_list('StudentID', flat=True).first()
    
    if student_id:
        try:
            student = Student.objects.get(StudentID=student_id)
            student_name = student.StudentName
        except Student.DoesNotExist:
            pass
        
        semesters = SemesterGrading.objects.filter(StudentID=student_id, Deleted=0).values_list('SemesterPeriodNumber', flat=True)
            
    context = {
        'student_name': student_name,
        'semesters': semesters,
    }
    return render(request, 'check_grades.html', context)


@login_required
@student_role_required
def get_grades(request):
    selected_semester = None
    course_info = []
    grade_sheet_data = {}
    semesters = []

    student_name = None
    student_id = Student.objects.filter(UserID=request.user.UserID).values_list('StudentID', flat=True).first()

    if student_id:
        try:
            student = Student.objects.get(StudentID=student_id)
            student_name = student.StudentName
        except Student.DoesNotExist:
            pass

        semesters = list(SemesterGrading.objects.filter(StudentID=student_id, Deleted=0).values_list('SemesterPeriodNumber', flat=True))
        semesters.sort(key=lambda x: int(x[1:]))

        if request.method == "POST":
            selected_semester = request.POST.get("selected_semester", semesters[0])
        else:
            selected_semester = semesters[0] if semesters else None

        if selected_semester:
            sem_grading_id = SemesterGrading.objects.filter(StudentID=student_id, SemesterPeriodNumber=selected_semester, Deleted=0).values_list('SemesterGradingID', flat=True).first()

            course_grading = CourseGrading.objects.filter(SemesterGradingID=sem_grading_id).values_list('CourseNumber', 'GradePointID')

            course_codes = [cg[0] for cg in course_grading]
            gp_ids = {gp[0]: gp[1] for gp in course_grading}

            course_info = Course.objects.filter(CourseNumber__in=course_codes).values('CourseNumber', 'CourseName', 'CourseCredits')

            grades = GradePoint.objects.filter(GradePointID__in=gp_ids.values()).values('GradePointID', 'Grade')

            grade_dict = {grade['GradePointID']: grade['Grade'] for grade in grades}

            for course in course_info:
                course['Grade'] = grade_dict.get(gp_ids[course['CourseNumber']], 'N/A')

            sem_id = SemesterGrading.objects.filter(StudentID=student_id, SemesterPeriodNumber=selected_semester, Deleted=0).values_list('SemesterID', flat=True).first()

            grade_sheet_data = GradeSheetData.objects.filter(StudentID=student_id, SemesterID=sem_id).values(
                'SemesterCredits', 'AccumulatedCredits', 'SGPA', 'CGPA').first()

            grade_points_data = GradePoint.objects.all().values('Grade', 'GradePointValue')
            for gradepoint in grade_points_data:
                if gradepoint["Grade"] == 'W' or gradepoint["Grade"] == 'RC':
                    gradepoint["GradePointValue"] = 0
                elif gradepoint["Grade"] == 'S' or gradepoint["Grade"] == 'X' or gradepoint["Grade"] == 'I':
                    gradepoint["GradePointValue"] = '-'

            current_index = semesters.index(selected_semester)
            prev_semester = semesters[current_index - 1] if current_index > 0 else None
            next_semester = semesters[current_index + 1] if current_index < len(semesters) - 1 else None

            context = {
                'student_name': student_name,
                'course_info': course_info,
                'grade_sheet_data': grade_sheet_data,
                'grade_points_data': grade_points_data,
                'semesters': semesters,
                'selected_semester': selected_semester,
                'prev_semester': prev_semester,
                'next_semester': next_semester,
            }
            return render(request, 'get_grades.html', context)

    return redirect('check_grades')



@login_required
@super_admin_required
def create_batch(request):
    if request.method == 'POST':
        batch_year = request.POST.get('batch_year')
        total_students = request.POST.get('total_students')
        batch_description = request.POST.get('batch_description')

        # Generate the next batch ID
        existing_batches = Batch.objects.all().order_by('BatchID')
        if existing_batches:
            last_batch_id = existing_batches.last().BatchID
            new_batch_id = 'B{:03}'.format(int(last_batch_id[1:]) + 1)
        else:
            new_batch_id = 'B001'

        try:
            # Create the new batch
            new_batch = Batch(BatchID=new_batch_id, BatchYear=batch_year, TotalStudent=total_students, BatchDescription=batch_description)
            new_batch.save()

            # Display success message
            messages.success(request, 'New batch created successfully!')

            # Redirect to the user management page
            return redirect('create_batch')
        except Exception as e:
            # Display error message
            messages.error(request, f'Error creating new batch: {str(e)}')

    current_year = datetime.datetime.now().year
    context = {
        'current_year': current_year,
    }
    return render(request, 'create_batch.html', context)


@login_required
@super_admin_required
def manage_roles(request, email=None):
    if request.method == 'POST':
        email = request.POST.get('email')
        
    assigned_roles_list = None
    available_roles_list = None
    selected_user = None
    selected_user_type = None

    if email:
        try:
            user_id = MIITUsers.objects.filter(username=email).values_list('UserID', flat=True).first()
            if not user_id:
                raise MIITUsers.DoesNotExist

            try:
                selected_user = Student.objects.get(UserID=user_id)
                selected_user_type = 'Student'
            except Student.DoesNotExist:
                try:
                    selected_user = FacultyStaff.objects.get(UserID=user_id)
                    selected_user_type = 'Faculty'
                except FacultyStaff.DoesNotExist:
                    messages.error(request, 'No user found with this email.', extra_tags='not_found')

            if selected_user:
                user_roles = MIITUserRole.objects.filter(UserID=user_id)

                # Get the RoleIDs from the MIITUserRole queryset
                role_ids = user_roles.values_list('RoleID', flat=True)

                # Query the MIITRole table to get the role descriptions for these RoleIDs
                assigned_roles = MIITRole.objects.filter(RoleID__in=role_ids)

                # To fetch the results as a list of dictionaries (if needed)
                assigned_roles_list = list(assigned_roles.values('RoleID', 'RoleDescription'))

                user_assigned_roles = MIITUserRole.objects.filter(UserID=user_id).values_list('RoleID', flat=True)

                # Query the MIITRole table to get roles where RoleID is not in the user_assigned_roles
                available_roles = MIITRole.objects.exclude(RoleID__in=user_assigned_roles)

                # To fetch the results as a list of dictionaries (if needed)
                available_roles_list = list(available_roles.values('RoleID', 'RoleDescription'))  
        
        except MIITUsers.DoesNotExist:
            messages.error(request, 'No user found with this email.', extra_tags='not_found')
    else:
        messages.error(request, 'User not found.')

    if 'add_role' in request.POST:
        email = request.POST.get('email')
        print(email)
        user_id = MIITUsers.objects.filter(username=email).values_list('UserID', flat=True).first()
        role_id = request.POST.get('add_role')
        try:
            role = MIITRole.objects.get(RoleID=role_id)  # Retrieve the MIITRole instance
            MIITUserRole.objects.create(UserID_id=user_id, RoleID=role)  # Use the instance directly
            return redirect('manage_roles_with_email', email=email)
        except MIITRole.DoesNotExist:
            messages.error(request, 'Role does not exist.', extra_tags='email_not_found')

    if 'delete_role' in request.POST:
        email = request.POST.get('email')
        user_id = MIITUsers.objects.filter(username=email).values_list('UserID', flat=True).first()
        role_id = request.POST.get('delete_role')
        MIITUserRole.objects.filter(UserID=user_id, RoleID=role_id).delete()
        return redirect('manage_roles_with_email', email=email)

    context = {
        'assigned_roles': assigned_roles_list,
        'available_roles': available_roles_list,
        'selected_user': selected_user,
        'email': email,
        'selected_user_type': selected_user_type,
    }
    return render(request, 'manage_roles.html', context)


@login_required
@super_admin_required
def edit_user_account(request, email=None):
    if request.method == 'POST':
        email = request.POST.get('email', None)
    selected_user = None
    user_name = None
    user = None
    selected_user_type = None
    assigned_roles_list = []
    available_roles_list = []

    if email:
        try:
            user = MIITUsers.objects.filter(username=email).first()
            if not user:
                raise MIITUsers.DoesNotExist

            try:
                selected_user = Student.objects.filter(UserID=user.UserID).values('UserID').first()
                user_name = Student.objects.get(UserID=user.UserID)
                selected_user_type = 'Student'
            except Student.DoesNotExist:
                try:
                    selected_user = FacultyStaff.objects.filter(UserID=user.UserID).values('UserID').first()
                    user_name = FacultyStaff.objects.get(UserID=user.UserID)
                    selected_user_type = 'Faculty'
                except FacultyStaff.DoesNotExist:
                    messages.error(request, 'No user found with this email.', extra_tags='not_found')

            if selected_user:
                user_roles = MIITUserRole.objects.filter(UserID=user.UserID)
                assigned_roles = MIITRole.objects.filter(RoleID__in=user_roles.values_list('RoleID', flat=True))
                assigned_roles_list = list(assigned_roles.values('RoleID', 'RoleDescription'))
                available_roles = MIITRole.objects.exclude(RoleID__in=user_roles.values_list('RoleID', flat=True))
                available_roles_list = list(available_roles.values('RoleID', 'RoleDescription'))

        except MIITUsers.DoesNotExist:
            messages.error(request, 'No user found with this email.', extra_tags='not_found')

    if request.method == 'POST' and 'update_user' in request.POST:
        user_id = request.POST.get('user_id')
        user = get_object_or_404(MIITUsers, UserID=user_id)
        email = request.POST.get('new_email')
        print(user)
        user.username = email
        user.UserStatus = request.POST.get('status')
        user.save()
        messages.success(request, 'User information updated successfully.', extra_tags='update_user')
        return redirect('edit_user_account_with_email', email)

    if request.method == 'POST' and 'reset_password' in request.POST:
        user_id = request.POST.get('user_id')
        user = get_object_or_404(MIITUsers, UserID=user_id)
        new_password = 'Thanks123!'
        user.set_password(new_password)
        user.save()
        messages.success(request, f'Password reset successfully. New password is: {new_password}', extra_tags='reset_password')
        return redirect('edit_user_account_with_email', user.username)

    if request.method == 'POST' and 'add_role' in request.POST:
        user_id = request.POST.get('user_id')
        user = get_object_or_404(MIITUsers, UserID=user_id)
        role_id = request.POST.get('add_role')
        try:
            role = MIITRole.objects.get(RoleID=role_id)
            MIITUserRole.objects.create(UserID_id=user_id, RoleID=role)
            messages.success(request, 'Role added successfully.', extra_tags='add_role')
            return redirect('edit_user_account_with_email', user.username)
        except MIITRole.DoesNotExist:
            messages.error(request, 'Role does not exist.', extra_tags='role_not_found')

    if request.method == 'POST' and 'delete_role' in request.POST:
        user_id = request.POST.get('user_id')
        user = get_object_or_404(MIITUsers, UserID=user_id)
        role_id = request.POST.get('delete_role')
        MIITUserRole.objects.filter(UserID=user_id, RoleID=role_id).delete()
        messages.success(request, 'Role removed successfully.', extra_tags='delete_role')
        return redirect('edit_user_account_with_email', user.username)
    
    if request.method == 'POST' and 'security' in request.POST:
        user_id = request.POST.get('user_id')
        user = get_object_or_404(MIITUsers, UserID=user_id)
        user.mfa_enabled = request.POST.get('mfa')
        user.acc_locked = request.POST.get('locked')
        user.save()
        messages.success(request, 'Security settings updated successfully.', extra_tags='security')
        return redirect('edit_user_account_with_email', email=user.username)

    
    if request.method == 'POST' and 'delete_account' in request.POST:
        user_id = request.POST.get('user_id')
        user = get_object_or_404(MIITUsers, UserID=user_id)
        user.Deleted = True
        user.save()
        messages.success(request, f'User account ({user_id}) deleted successfully.', extra_tags='delete_account')
        return redirect('edit_user_account_with_email', user.username)

    context = {
        'selected_user': selected_user,
        'user_name': user_name,
        'selected_user_type': selected_user_type,
        'assigned_roles': assigned_roles_list,
        'available_roles': available_roles_list,
        'user': user,
    }
    return render(request, 'edit_user_account.html', context)


@login_required
@super_admin_required
def edit_user_information(request, email=None):
    if request.method == 'POST':
        email = request.POST.get('email')
    selected_user = None
    selected_user_type = None
    discipline = None
    program = None
    batch = None
    user_id = None

    if email:
        try:
            user_id = MIITUsers.objects.filter(username=email).values_list('UserID', flat=True).first()
            if not user_id:
                raise MIITUsers.DoesNotExist

            try:
            
                selected_user = Student.objects.get(UserID=user_id)
                selected_user_type = 'Student'
                discipline = selected_user.DisciplineID
                program = selected_user.ProgramID
                batch = selected_user.BatchID

            except Student.DoesNotExist:
                try:
                    selected_user = FacultyStaff.objects.get(UserID=user_id)
                    selected_user_type = 'Faculty'
                except FacultyStaff.DoesNotExist:
                    messages.error(request, 'No user found with this email.', extra_tags='not_found')

        except MIITUsers.DoesNotExist:
            messages.error(request, 'No user found with this email.', extra_tags='not_found')

    context = {
        'user_id':user_id,
        'selected_user': selected_user,
        'selected_user_type': selected_user_type,
        'email': email,
        'discipline': discipline,
        'program': program,
        'batch': batch,
        'disciplines': Discipline.objects.all(),
        'programs': Program.objects.all(),
        'batches': Batch.objects.all(),
    }
    return render(request, 'edit_user_information.html', context)


@login_required
@super_admin_required
@csrf_exempt
def manage_discipline(request):
    if request.method == 'GET':
        disciplines = Discipline.objects.all()
        context = {
            'disciplines': disciplines,
            'messages': [],
        }
        return render(request, 'manage_discipline.html', context)

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)

        if 'create_discipline' in data:
            full_name = data.get('DisciplineFullName')
            short_name = data.get('DisciplineShortName')
            description = data.get('DisciplineDescription')

            if not all([full_name, short_name, description]):
                return JsonResponse({'success': False, 'message': 'All fields are required for creating a new discipline.'}, status=400)

            discipline_count = Discipline.objects.count() + 1
            discipline_id = f"D{discipline_count:02d}"

            try:
                Discipline.objects.create(
                    DisciplineID=discipline_id,
                    DisciplineFullName=full_name,
                    DisciplineShortName=short_name,
                    DisciplineDescription=description
                )
                return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)

        elif 'edit_discipline' in data:
            discipline_id = data.get('DisciplineID')
            if not discipline_id:
                return JsonResponse({'success': False, 'message': 'DisciplineID is required'}, status=400)

            discipline = get_object_or_404(Discipline, DisciplineID=discipline_id)
            discipline.DisciplineFullName = data.get('DisciplineFullName', discipline.DisciplineFullName)
            discipline.DisciplineShortName = data.get('DisciplineShortName', discipline.DisciplineShortName)
            discipline.DisciplineDescription = data.get('DisciplineDescription', discipline.DisciplineDescription)
            discipline.save()
            return JsonResponse({'success': True})

        elif 'delete_discipline' in data:
            discipline_id = data.get('DisciplineID')
            if not discipline_id:
                return JsonResponse({'success': False, 'message': 'DisciplineID is required'}, status=400)

            discipline = get_object_or_404(Discipline, DisciplineID=discipline_id)
            discipline.delete()
            return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)

@login_required
def registration_offered_courses(request):
    student = get_student(request.user.UserID)
    if request.method == 'POST':
        try:
            courses = request.POST.getlist('courses')
            roll_number = student.RollNumber 
            registrationtype = 'As per chart'
            semester = Semester.objects.filter(SemesterStatus='Current').values_list('SemesterID' ,flat=True)
            print(semester)
            for course in courses:
                SemesterRegistrationData.objects.create(
                    RollNumber=roll_number,
                    CourseNumber=course,
                    RegistrationType= registrationtype,
                    SemesterID= semester,  
                    CourseRegistered=course,
                    RegistrationStatus=1
                )
            # StudentToRegisterData.RegistrationTag='R';
            messages.success(request, 'Courses Registered successfully!')
        except Exception as e:
            messages.error(request, f'Error during course registration: {str(e)}')
        return redirect('registration_offered_courses')

    base_chart_id = get_base_chart_id(student)
    completed_courses = get_completed_courses(student)
    next_courses_info = get_next_courses_info(base_chart_id, completed_courses)

    context = {
        'next_semester': next_courses_info['next_semester'],
        'student_name': student.StudentName,
        'courses': next_courses_info['courses']
    }

    return render(request, 'registration_offered_courses.html', context)

def get_student(user_id):
    return Student.objects.get(UserID=user_id)

def get_base_chart_id(student):
    batch_id = Batch.objects.filter(BatchID=student.BatchID).values('BatchID')
    base_chart_id = BaseChart.objects.filter(DisciplineID=student.DisciplineID, ProgramID=student.ProgramID).values('BaseChartID')
    return base_chart_id

def get_completed_courses(student):
    semester_grading_ids = SemesterGrading.objects.filter(StudentID=student.StudentID).values('SemesterGradingID')
    completed_courses_ = CourseGrading.objects.filter(SemesterGradingID__in=Subquery(semester_grading_ids)).values('CourseNumber')
    return [course['CourseNumber'] for course in completed_courses_]

def get_next_courses_info(base_chart_id, completed_courses):
    if not base_chart_id.exists():
        return {'next_semester': "Done!", 'courses': []}

    base_chart_id_value = base_chart_id[0]['BaseChartID']
    all_course_codes = OfferedCourses.objects.filter(BaseChartID=base_chart_id_value).values('CourseNumber', 'SemesterPeriodNumber', 'YearNumber')
    all_courses_df = pd.DataFrame(list(all_course_codes))
    next_courses = all_courses_df[~all_courses_df['CourseNumber'].isin(completed_courses)]

    if not next_courses.empty:
        next_course = next_courses.iloc[0]
        next_semester_number = next_course['SemesterPeriodNumber']
        course_codes = OfferedCourses.objects.filter(
            SemesterPeriodNumber=next_semester_number, BaseChartID=base_chart_id_value
        ).values('CourseNumber')
        courses = Course.objects.filter(
            CourseNumber__in=Subquery(course_codes)
        ).values('CourseNumber', 'CourseName', 'CourseCredits')
        return {'next_semester': next_semester_number, 'courses': courses}
    else:
        return {'next_semester': "Done!", 'courses': []}


@login_required
def completed_courses(request):
    student_name = None
    semester_courses = defaultdict(list)

    # Get the student information
    roll_number = Student.objects.filter(UserID=request.user.UserID).values_list('RollNumber', flat=True).first()

    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT StudentName FROM student WHERE RollNumber = %s", [roll_number])
            student_name = cursor.fetchone()[0]
    except (Student.DoesNotExist, TypeError):
        pass

    # Get the semester grades for the student
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT sg.SemesterPeriodNumber, c.CourseNumber, c.CourseName, c.CourseCredits, c.CourseDescription, gp.Grade
            FROM semestergrading sg
            JOIN coursegrading cg ON sg.SemesterGradingID = cg.SemesterGradingID
            JOIN course c ON cg.CourseNumber = c.CourseNumber
            JOIN gradepoint gp ON cg.GradePointID = gp.GradePointID
            JOIN student st ON sg.StudentID = st.StudentID
            WHERE st.RollNumber = %s AND sg.Deleted = 0 
        """, [roll_number])

        rows = cursor.fetchall()
        for row in rows:
            semester_number, course_code, course_name, course_credits, course_description, grade= row 
            semester_courses[semester_number].append({
                'CourseNumber': course_code,
                'CourseName': course_name,
                'CourseCredits': course_credits,
                'CourseDescription': course_description,
                'Grade':grade,
            })

    semester_courses = dict(semester_courses)

    context = {
        'student_name': student_name,
        'semester_courses': semester_courses,
    }

    return render(request, 'completed_courses.html', context)



@login_required
@student_role_required
def projects(request):
    user = request.user
    project = defaultdict(list)

    student_id = Student.objects.filter(UserID=user.UserID).values_list('StudentID', flat=True).first()
    subquery = StudentProjectRegistration.objects.filter(StudentID=student_id).values('ProjectNumber')
    project = Project.objects.filter(ProjectNumber__in=Subquery(subquery))

    program_id = Student.objects.filter(UserID=user.UserID).values_list('ProgramID', flat=True).first()
    current_semesters = Semester.objects.filter(SemesterStatus='Current')
    offered_projects = Project.objects.filter(ProgramID=program_id, SemesterID__in=current_semesters, Deleted=False)

    context = {
        'project': project,
        'offered_projects':offered_projects,
        'projects': Project.objects.all(),
        'user': user, 
        'messages': messages.get_messages(request), 
    }
    return render(request, 'projects.html', context)



def tech_news(request):
    fetch_news_articles()
    news_articles = News.objects.order_by('-published_at')[:50]  # Display the latest 10 articles
    return render(request, 'tech_news.html', {'news_articles': news_articles})



@login_required
@student_affairs_role_required
def student_affairs_dashboard(request):
    return render(request, 'student_affairs_dashboard.html')


@login_required
@student_affairs_super_admin_required
def generate_transcript(request):
    email = None
    selected_user = None
    if request.method == 'POST':
        email = request.POST.get('email')

        try:
            user_id = MIITUsers.objects.filter(username=email).values_list('UserID', flat=True).first()
            if not user_id:
                raise MIITUsers.DoesNotExist
            
            selected_user = Student.objects.filter(UserID=user_id).values().first()
            
            if not selected_user:
                raise Student.DoesNotExist

            if 'generate_transcript' in request.POST:
                student_id = Student.objects.filter(UserID=user_id).values_list('StudentID', flat=True).first()
                student = Student.objects.get(StudentID=student_id)
                student_name = student.StudentName
                roll_number = student.RollNumber
                program = Program.objects.filter(ProgramID=student.ProgramID).values_list('ProgramFullName', flat=True).first()
                discipline = Discipline.objects.filter(DisciplineID=student.DisciplineID).values_list('DisciplineFullName', flat=True).first()
                
                # Fetch semesters and group by academic year
                semesters = SemesterGrading.objects.filter(StudentID=student_id, Deleted=0).values('SemesterGradingID', 'SemesterID', 'SemesterPeriodNumber')
                semesters = sorted(semesters, key=lambda x: int(x['SemesterPeriodNumber'][1:]))  # Sort by semester number
                
                academic_years = {}
                for sem in semesters:
                    semester_id = sem['SemesterID']
                    academic_year_id = Semester.objects.get(SemesterID=semester_id).AY_ID
                    if academic_year_id not in academic_years:
                        academic_years[academic_year_id] = {
                            'academic_year_label': f"Academic Year {academic_year_id}",
                            'semesters': []
                        }
                    sem['course_info'] = CourseGrading.objects.filter(SemesterGradingID=sem['SemesterGradingID']).values('CourseNumber', 'GradePointID')
                    for course in sem['course_info']:
                        course_info = Course.objects.get(CourseNumber=course['CourseNumber'])
                        grade = GradePoint.objects.get(GradePointID=course['GradePointID']).Grade
                        course['CourseName'] = course_info.CourseName
                        course['CourseCredits'] = course_info.CourseCredits
                        course['Grade'] = grade
                    
                    grade_sheet = GradeSheetData.objects.filter(StudentID=student_id, SemesterID=semester_id).values_list('SemesterCredits', 'AccumulatedCredits', 'SGPA', 'CGPA').first()
                    sem['grade_sheet'] = grade_sheet
                    academic_years[academic_year_id]['semesters'].append(sem)

                grade_points_data = GradePoint.objects.filter(~Q(Grade='SNO') & ~Q(Grade='DNR')).values('Grade', 'GradePointValue')

                # Create the HttpResponse object with the appropriate PDF headers.
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="transcript_{student_name}.pdf"'
                
                buffer = BytesIO()
                p = canvas.Canvas(buffer, pagesize=A4)
                width, height = A4
                p.setPageSize((height, width))  # Landscape mode
                
                def draw_header(page_num):
                    logo_path = os.path.join(settings.STATICFILES_DIRS[0], 'images', 'miit-logo.png')
                    p.drawImage(logo_path, 50, width - 80, width=60, height=60, mask='auto')
                    p.setFont("Helvetica-Bold", 12)
                    p.drawString(width - 60, width - 40, "Myanmar Institute of Information Technology")
                    p.drawString(width - 60 , width - 57, "Mandalay, Myanmar")
                    p.line(40, width - 90, height - 40, width - 90)
                    p.setFont("Helvetica-Bold", 12)
                    p.drawString((height / 2) - 50, width - 110, "Academic Transcript")
                    p.setFont("Helvetica-Bold", 10)
                    p.drawString(50, width - 120, "Name: ")
                    p.setFont("Helvetica", 10)
                    p.drawString(130, width - 120, f"{student_name}")
                    p.setFont("Helvetica-Bold", 10)
                    p.drawString(50, width - 135, f"Roll Number:")
                    p.setFont("Helvetica", 10)
                    p.drawString(130, width - 135, f"{roll_number}")

                    p.setFont("Helvetica-Bold", 10)
                    p.drawString(width - 50, width - 120, f"Program:")
                    p.setFont("Helvetica", 10)
                    p.drawString(width + 20, width - 120, f"{program}")
                    p.setFont("Helvetica-Bold", 10)
                    p.drawString(width - 50, width - 135, f"Discipline:")
                    p.setFont("Helvetica", 10)
                    p.drawString(width + 20, width - 135, f"{discipline}")
                
                def draw_footer(page_num):
                    p.line(40, width - 520, height - 40, width - 520)
                    p.setFont("Helvetica-Bold", 10)
                    p.drawString(40, 50 , "Copyright MIIT")
                    p.drawRightString(height - 40, 50 , f"Page {page_num}")
                    p.setFont("Helvetica", 6)
                    # Prepare data for column layout
                    grades = [item['Grade'] for item in grade_points_data]
                    grade_points = [item['GradePointValue'] for item in grade_points_data]
                    data = [["Grade"] + grades, ["Grade Point"] + grade_points]

                    col_widths = [60] + [30] * (len(grades) - 1)
                    table = Table(data, colWidths=col_widths)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (0, -1), colors.white),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONT', (0, 0), (-1, -1), 'Helvetica', 6),
                        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 6),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  # Adjusted bottom padding
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))
                    table.wrapOn(p, height - 80, width - 100)
                    table.drawOn(p, 200, 30)


                def draw_table_for_semesters(semester, x, y):
                    col_widths = [50, 100, 30, 30]  # Manually set column widths

                    p.setFont("Helvetica", 10)
                    p.drawString(x, y+20, f"Semester {semester['SemesterPeriodNumber']}")
                    y -= 10

                    # Create grade table
                    grade_data = [["Course Code", "Course Name", "Credits", "Grade"]]
                    for course in semester['course_info']:
                        grade_data.append([course['CourseNumber'], course['CourseName'], course['CourseCredits'], course['Grade']])

                    table = Table(grade_data, colWidths=col_widths)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 7),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONT', (0, 1), (-1, -1), 'Helvetica', 6),
                    ]))
                    table.wrapOn(p, height - 80, width - 20)
                    table.drawOn(p, x, y - 120)
                    y -= 180

                    # Semester Credits Table
                    sem_credits_data = [
                        ["Semester Credits", semester['grade_sheet'][0]],
                        ["Accumulated Credits", semester['grade_sheet'][1]],
                    ]
                    sem_credits_col_widths = [70, 30]  # Manually set column widths
                    sem_credits_table = Table(sem_credits_data, colWidths=sem_credits_col_widths)
                    sem_credits_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (0, -1), colors.white),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONT', (0, 0), (-1, -1), 'Helvetica', 6),
                        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 6),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  # Adjusted bottom padding
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))
                    sem_credits_table.wrapOn(p, height - 80, width - 200)
                    sem_credits_table.drawOn(p, x, y - (len(sem_credits_data) * 15))

                    # SGPA/CGPA Table
                    sgpa_cgpa_data = [
                        ["SGPA", semester['grade_sheet'][2]],
                        ["CGPA", semester['grade_sheet'][3]],
                    ]
                    sgpa_cgpa_col_widths = [40, 30]  # Manually set column widths
                    sgpa_cgpa_table = Table(sgpa_cgpa_data, colWidths=sgpa_cgpa_col_widths)
                    sgpa_cgpa_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (0, -1), colors.white),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONT', (0, 0), (-1, -1), 'Helvetica', 6),
                        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 6),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  # Adjusted bottom padding
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))
                    sgpa_cgpa_table.wrapOn(p, height - 80, width - 200)
                    sgpa_cgpa_table.drawOn(p, x + 140, y - (len(sgpa_cgpa_data) * 15))
                    y -= len(sgpa_cgpa_data) * 20 + 20

                    

                page_num = 1
                
                for academic_year_id, academic_year_data in academic_years.items():
                    draw_header(page_num)
                    draw_footer(page_num)
                    y = width - 160

                    p.setFont("Helvetica-Bold", 10)
                    p.drawString(50, y, academic_year_data['academic_year_label'])
                    y -= 40
                    x=40
                    # Draw the tables for the academic year
                    for semester in academic_year_data['semesters']:
                        draw_table_for_semesters(semester, x, y)
                        x += 280
                    page_num += 1
                    p.showPage()
                    y -= 40  # Space between different academic years

                p.save()

                # Get the value of the BytesIO buffer and write it to the response.
                pdf = buffer.getvalue()
                buffer.close()
                response.write(pdf)
                return response
        
        except MIITUsers.DoesNotExist:
            print("hello")
            messages.error(request, 'No user found with this email.', extra_tags='user_not_found')

        except Student.DoesNotExist:
            messages.error(request, 'Student not found.', extra_tags='student_not_found') 

    context = {
        'email':email,
        'selected_user':selected_user
    }

    if request.session.get('selected_role') == 'R07':
        return render(request, 'generate_transcript_sa.html', context)
    elif request.session.get('selected_role') in ['R01','R03'] :
        return render(request, 'generate_transcript_admin.html', context)

@login_required
@student_affairs_role_required
def edit_user_status(request, email=None):
    if request.method == 'POST':
        email = request.POST.get('email', None)
    selected_user = None
    user_name = None
    user = None
    selected_user_type = None

    if email:
        try:
            user = MIITUsers.objects.filter(username=email).first()
            if not user:
                raise MIITUsers.DoesNotExist

            try:
                selected_user = Student.objects.filter(UserID=user.UserID).values('UserID').first()
                user_name = Student.objects.get(UserID=user.UserID)
                selected_user_type = 'Student'
            except Student.DoesNotExist:
                try:
                    selected_user = FacultyStaff.objects.filter(UserID=user.UserID).values('UserID').first()
                    user_name = FacultyStaff.objects.get(UserID=user.UserID)
                    selected_user_type = 'Faculty'
                except FacultyStaff.DoesNotExist:
                    messages.error(request, 'No user found with this email.', extra_tags='not_found')

            if selected_user:
                user_roles = MIITUserRole.objects.filter(UserID=user.UserID)

        except MIITUsers.DoesNotExist:
            messages.error(request, 'No user found with this email.', extra_tags='not_found')

    if request.method == 'POST' and 'update_user' in request.POST:
        user_id = request.POST.get('user_id')
        user = get_object_or_404(MIITUsers, UserID=user_id)
        email = request.POST.get('new_email')
        print(user)
        user.username = email
        user.UserStatus = request.POST.get('status')
        user.save()
        messages.success(request, 'User account status updated successfully.', extra_tags='update_user')
        return redirect('edit_user_status_with_email', email)

    context = {
        'selected_user': selected_user,
        'user_name': user_name,
        'selected_user_type': selected_user_type,
        'user': user,
    }
    return render(request, 'edit_user_status.html', context)


@login_required
@student_affairs_super_admin_required
def information_center(request, email=None):
    if request.method == 'POST':
        email = request.POST.get('email')
    selected_user = None
    selected_user_type = None
    discipline = None
    program = None
    batch = None
    user_id = None

    if email:
        try:
            user_id = MIITUsers.objects.filter(username=email).values_list('UserID', flat=True).first()
            if not user_id:
                raise MIITUsers.DoesNotExist

            try:
            
                selected_user = Student.objects.get(UserID=user_id)
                selected_user_type = 'Student'
                discipline = selected_user.DisciplineID
                program = selected_user.ProgramID
                batch = selected_user.BatchID

            except Student.DoesNotExist:
                try:
                    selected_user = FacultyStaff.objects.get(UserID=user_id)
                    selected_user_type = 'Faculty'
                except FacultyStaff.DoesNotExist:
                    messages.error(request, 'No user found with this email.', extra_tags='not_found')

        except MIITUsers.DoesNotExist:
            messages.error(request, 'No user found with this email.', extra_tags='not_found')

    context = {
        'user_id':user_id,
        'selected_user': selected_user,
        'selected_user_type': selected_user_type,
        'email': email,
        'discipline': discipline,
        'program': program,
        'batch': batch,
        'disciplines': Discipline.objects.all(),
        'programs': Program.objects.all(),
        'batches': Batch.objects.all(),
    }
    return render(request, 'information_center.html', context)


@login_required
def autocomplete_emails(request):
    if 'term' in request.GET:
        term = request.GET.get('term')
        emails = list(MIITUsers.objects.filter(username__istartswith=term).values_list('username', flat=True))
        return JsonResponse({'emails': emails}, safe=False)
    return JsonResponse([], safe=False)


@login_required
@student_affairs_super_admin_required
def update_student_personal(request):
    if 'update_student_info' in request.POST:
        img = request.FILES.get('avatar')

        student_name = request.POST.get('student_name')
        student_nrc = request.POST.get('student_nrc')
        nationality = request.POST.get('nationality')
        religion = request.POST.get('religion')
        student_dob = request.POST.get('student_dob')
        student_phone = request.POST.get('student_phone')
        student_email = request.POST.get('student_email')
        print(student_email)

        user_id = MIITUsers.objects.filter(username=student_email).values_list('UserID', flat=True).first()
        student_id = Student.objects.filter(UserID=user_id).values_list('StudentID', flat=True).first()

        # Update student information
        Student.objects.filter(StudentID=student_id).update(
            StudentName=student_name,
            StudentNRC=student_nrc,
            Nationality=nationality,
            Religion=religion,
            StudentDOB=student_dob,
            StudentPhone=student_phone
        )

        # Save the image
        if img:
            image_path = os.path.join(settings.STATICFILES_DIRS[0], 'images/users/Students', f'{user_id}.png')
            with open(image_path, 'wb+') as destination:
                for chunk in img.chunks():
                    destination.write(chunk)

        messages.success(request, 'Personal information updated successfully.', extra_tags='personal_info')

    if request.session.get('selected_role') == 'R07':
        return redirect('information_center_with_email', email=student_email)
    elif request.session.get('selected_role') in ['R01','R03'] :
        return redirect('edit_user_information_with_email', email=student_email)



@login_required
@student_affairs_super_admin_required
def update_miit_academic_info(request):
    student_email = None
    if 'update_academic_info' in request.POST:
        student_email = request.POST.get('student_email')
        rollnumber = request.POST.get('roll_number')
        program_id = request.POST.get('program_id')
        discipline_id = request.POST.get('discipline_id')
        batch_id = request.POST.get('batch_id')
        section = request.POST.get('section_name')
        miit_id = request.POST.get('miit_id')
        acb_status = 1 if request.POST.get('acb_status') else 0

        user_id = MIITUsers.objects.filter(username=student_email).values_list('UserID', flat=True).first()
        student_id = Student.objects.filter(UserID=user_id).values_list('StudentID', flat=True).first()

        # Update academic information
        Student.objects.filter(StudentID=student_id).update(
            RollNumber=rollnumber,
            ProgramID_id=program_id,
            DisciplineID_id=discipline_id,
            BatchID_id=batch_id,
            SectionName=section,
            MIITID=miit_id,
            ACBStatus=acb_status
        )

        messages.success(request, 'Academic information updated successfully.', extra_tags='academic_info')

    if request.session.get('selected_role') == 'R07':
        return redirect('information_center_with_email', email=student_email)
    elif request.session.get('selected_role') in ['R01','R03'] :
        return redirect('edit_user_information_with_email', email=student_email)




@login_required
@student_affairs_super_admin_required
def update_highschool(request):
    student_email = None
    if 'update_highschool' in request.POST:
        student_email = request.POST.get('student_email')
        matricrollnumber = request.POST.get('matric_roll_number')
        matricexamyear = request.POST.get('matric_exam_year')

        user_id = MIITUsers.objects.filter(username=student_email).values_list('UserID', flat=True).first()
        student_id = Student.objects.filter(UserID=user_id).values_list('StudentID', flat=True).first()

        # Update high school information
        Student.objects.filter(StudentID=student_id).update(
            MatricRollNumber=matricrollnumber,
            MatricExamYear=matricexamyear
        )

        messages.success(request, 'High school information updated successfully.', extra_tags='highschool_info')

    if request.session.get('selected_role') == 'R07':
        return redirect('information_center_with_email', email=student_email)
    elif request.session.get('selected_role') in ['R01','R03'] :
        return redirect('edit_user_information_with_email', email=student_email)



@login_required
@student_affairs_super_admin_required
def update_guardian_info(request):
    student_email = None
    if 'update_guardian_info' in request.POST:
        student_email = request.POST.get('student_email')
        father_name = request.POST.get('father_name')
        father_nrc = request.POST.get('father_nrc')
        father_phone = request.POST.get('father_phone')
        mother_name = request.POST.get('mother_name')
        mother_nrc = request.POST.get('mother_nrc')
        mother_phone = request.POST.get('mother_phone')

        user_id = MIITUsers.objects.filter(username=student_email).values_list('UserID', flat=True).first()
        student_id = Student.objects.filter(UserID=user_id).values_list('StudentID', flat=True).first()

        # Update guardian information
        Student.objects.filter(StudentID=student_id).update(
            FatherName = father_name,
            FatherNRC = father_nrc,
            FatherPhoneNumber = father_phone,
            MotherName = mother_name,
            MotherNRC = mother_nrc,
            MotherPhoneNumber = mother_phone,
        )

        messages.success(request, 'Guardian information updated successfully.', extra_tags='guardian_info')

    if request.session.get('selected_role') == 'R07':
        return redirect('information_center_with_email', email=student_email)
    elif request.session.get('selected_role') in ['R01','R03'] :
        return redirect('edit_user_information_with_email', email=student_email)



@login_required
@student_affairs_super_admin_required
def update_address_info(request):
    if 'update_address' in request.POST:
        student_email = request.POST.get('student_email')
        address = request.POST.get('address')

        user_id = MIITUsers.objects.filter(username=student_email).values_list('UserID', flat=True).first()
        student_id = Student.objects.filter(UserID=user_id).values_list('StudentID', flat=True).first()

        # Update address information
        Student.objects.filter(StudentID=student_id).update(
            Address=address,
        )

        messages.success(request, 'Address updated successfully.', extra_tags='address_info')

    if request.session.get('selected_role') == 'R07':
        return redirect('information_center_with_email', email=student_email)
    elif request.session.get('selected_role') in ['R01','R03'] :
        return redirect('edit_user_information_with_email', email=student_email)


@login_required
@student_affairs_super_admin_required
def update_faculty_personal(request):
    if 'update_faculty_info' in request.POST:
        img = request.FILES.get('avatar')

        faculty_name = request.POST.get('faculty_name')
        salutation = request.POST.get('salutation')
        short_name = request.POST.get('short_name')
        designation = request.POST.get('designation')
        department = request.POST.get('department')
        nrc = request.POST.get('nrc')
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        faculty_email = request.POST.get('faculty_email')
        print(faculty_email)

        user_id = MIITUsers.objects.filter(username=faculty_email).values_list('UserID', flat=True).first()
        faculty_id = FacultyStaff.objects.filter(UserID=user_id).values_list('FacultyStaffID', flat=True).first()

        # Update faculty information
        FacultyStaff.objects.filter(FacultyStaffID=faculty_id).update(
            FacultyStaffName=faculty_name,
            Salutation=salutation,
            ShortName=short_name,
            Designation=designation,
            Department=department,
            NRC=nrc,
            Phone=phone,
            Address=address
        )

        # Save the image
        if img:
            image_path = os.path.join(settings.STATICFILES_DIRS[0], 'images/users/FacultyStaffs', f'{user_id}.png')
            with open(image_path, 'wb+') as destination:
                for chunk in img.chunks():
                    destination.write(chunk)

        messages.success(request, 'Personal information updated successfully.', extra_tags='faculty_personal_info')

    if request.session.get('selected_role') == 'R07':
        return redirect('information_center_with_email', email=faculty_email)
    elif request.session.get('selected_role') in ['R01','R03'] :
        return redirect('edit_user_information_with_email', email=faculty_email)

@login_required
def academic_calendar(request):
    return render(request, 'academic_calendar.html')


def success_view(request):
  return render(request, 'successful.html') 

def invalid_view(request):
  return render(request, 'invalid.html')
                   


def test_page(request):
   return render(request, 'testing.html')

