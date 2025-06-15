import openpyxl
from openpyxl.styles import Font, Alignment

# Create a new workbook
wb = openpyxl.load_workbook('data.xlsx')

# Create a new sheet for Batch
ws_batch = wb.create_sheet("Batch Data")

# Set the column headings for Batch
batch_column_headings = ["Batch ID", "Batch Year", "Total Students", "Batch Description"]
for col_num, heading in enumerate(batch_column_headings, 1):
    ws_batch.cell(row=1, column=col_num, value=heading)

# Set font and alignment for column headings in Batch sheet
font = Font(bold=True)
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for cell in ws_batch[1]:
    cell.font = font
    cell.alignment = alignment

# Generate batch_data up to 2023 excluding 2020 batch
batch_data = []
batch_count = 1
for year in range(2014, 2024):
    if year == 2020:
        continue  # Exclude 2020 batch
    batch_id = f"B{batch_count:03d}"
    batch_data.append([batch_id, year, 120, f"This is {year} batch."])
    batch_count += 1

# Insert the data into the Batch worksheet
for row_num, row_data in enumerate(batch_data, 2):
    for col_num, value in enumerate(row_data, 1):
        ws_batch.cell(row=row_num, column=col_num, value=value)

# Adjust column widths for Batch sheet
for column in ws_batch.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_batch.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

# Save the updated workbook
wb.save('data.xlsx')

print("Data instances for the Batch table generated successfully in data.xlsx")
