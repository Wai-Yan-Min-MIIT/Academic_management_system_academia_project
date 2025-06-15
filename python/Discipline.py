import openpyxl
from openpyxl.styles import Font, Alignment

# Create a new workbook
wb = openpyxl.load_workbook('data.xlsx')

# Create a new sheet for Discipline
ws_discipline = wb.create_sheet("Discipline Data")

# Set the column headings for Discipline
discipline_column_headings = ["Discipline ID", "Discipline Full Name", "Discipline Short Name", "Discipline Description"]
for col_num, heading in enumerate(discipline_column_headings, 1):
    ws_discipline.cell(row=1, column=col_num, value=heading)

# Set font and alignment for column headings in Discipline sheet
font = Font(bold=True)
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for cell in ws_discipline[1]:
    cell.font = font
    cell.alignment = alignment

# Insert the data instances for Discipline
discipline_data = [
    ['D01', 'Postgraduate Diploma in Software Development', 'PGDSD', '(PGDSD) programme'],
    ['D02', 'Computer Science and Engineering', 'CSE', 'Computer Science and Engineering (CSE)'],
    ['D03', 'Electronic Communication and Engineering', 'ECE', 'Electronic Communication and Engineering (ECE)'],
]

# Insert the data into the Discipline worksheet
for row_num, row_data in enumerate(discipline_data, 2):
    for col_num, value in enumerate(row_data, 1):
        ws_discipline.cell(row=row_num, column=col_num, value=value)

# Adjust column widths for Discipline sheet
for column in ws_discipline.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_discipline.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

# Save the updated workbook
wb.save('data.xlsx')

print("Data instances for the Discipline table generated successfully in data.xlsx")
