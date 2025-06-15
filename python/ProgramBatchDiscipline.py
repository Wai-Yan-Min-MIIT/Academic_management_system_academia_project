import openpyxl
from openpyxl.styles import Font, Alignment

# Function to generate sequential PBD_ID
def generate_pbd_id(count):
    return f"PBD{count:03d}"

# Create a new workbook
wb = openpyxl.load_workbook('data.xlsx')

# Create a new sheet for Discipline
ws_pbd = wb.create_sheet("ProgramBatchDiscipline Data")

# Set the column headings for ProgramBatchDiscipline
pbd_column_headings = ["PBD ID", "Program ID", "Batch ID", "Discipline ID"]
for col_num, heading in enumerate(pbd_column_headings, 1):
    ws_pbd.cell(row=1, column=col_num, value=heading)

# Set font and alignment for column headings in ProgramBatchDiscipline sheet
font = Font(bold=True)
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for cell in ws_pbd[1]:
    cell.font = font
    cell.alignment = alignment

# Generate ProgramBatchDiscipline data based on associations
pbd_data = []
pbd_count = 1

# 2014 batch has only D01 and P01
for _ in range(2):
    pbd_data.append([generate_pbd_id(pbd_count), 'P01', 'B001', 'D01'])
    pbd_count += 1

# 2015 batch has D01, D02, D03
for _ in range(2):
    pbd_data.append([generate_pbd_id(pbd_count), 'P01', 'B002', 'D01'])
    pbd_count += 1
for i in range(2):
    pbd_data.append([generate_pbd_id(pbd_count), 'P02', 'B002', f'D0{i + 2}'])
    pbd_count += 1

# 2016 batch has D01, D02, D03
for _ in range(2):
    pbd_data.append([generate_pbd_id(pbd_count), 'P01', 'B003', 'D01'])
    pbd_count += 1
for i in range(2):
    pbd_data.append([generate_pbd_id(pbd_count), 'P02', 'B003', f'D0{i + 2}'])
    pbd_count += 1

# Starting from 2017 batch, all batches have D02, D03 (excluding D01)
for batch_id in ['B004', 'B005', 'B006', 'B007', 'B008', 'B009']:
    for _ in range(2):
        pbd_data.append([generate_pbd_id(pbd_count), 'P02', batch_id, 'D02'])
        pbd_count += 1
    for _ in range(2):
        pbd_data.append([generate_pbd_id(pbd_count), 'P02', batch_id, 'D03'])
        pbd_count += 1

# Remove duplicate rows based on Program ID, Batch ID, and Discipline ID
unique_pbd_data = []
seen_combinations = set()
for row_data in pbd_data:
    program_id, batch_id, discipline_id = row_data[1], row_data[2], row_data[3]
    combination = (program_id, batch_id, discipline_id)
    if combination not in seen_combinations:
        unique_pbd_data.append(row_data)
        seen_combinations.add(combination)

# Update PBD_ID to be sequential
for idx, row_data in enumerate(unique_pbd_data, start=1):
    row_data[0] = generate_pbd_id(idx)

# Insert the data into the ProgramBatchDiscipline worksheet
for row_num, row_data in enumerate(unique_pbd_data, 2):
    for col_num, value in enumerate(row_data, 1):
        ws_pbd.cell(row=row_num, column=col_num, value=value)

# Adjust column widths for ProgramBatchDiscipline sheet
for column in ws_pbd.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_pbd.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

# Save the updated workbook
wb.save('data.xlsx')

print("Data instances for the ProgramBatchDiscipline table generated successfully in data.xlsx")
