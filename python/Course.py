import openpyxl

def process_course_data():
    # Load the SampleData2.xlsx workbook
    sample_data_workbook = openpyxl.load_workbook('SampleData2.xlsx')  # Update the path accordingly
    course_sheet = sample_data_workbook['Course']  # Assuming the sheet name is 'Course'

    # Load the data.xlsx workbook, or create it if it doesn't exist
    try:
        data_workbook = openpyxl.load_workbook('data.xlsx')
    except FileNotFoundError:
        data_workbook = openpyxl.Workbook()
    
    course_data_sheet = data_workbook.create_sheet("Course Data")

    # Define column headings for the new sheet
    headings = ["CourseCode", "ProgramID", "CourseName", "CourseCredits", "Deleted", "CourseDescription"]
    course_data_sheet.append(headings)

    # Process each row in the Course sheet from SampleData2.xlsx
    for row in course_sheet.iter_rows(min_row=2, values_only=True):
        course_code = row[0]  # Assuming CourseNumber/CourseCode is the first column
        if course_code is not None:
            program_id = "P01" if "PGDSD" in course_code or "PGDSDM" in course_code or "SD" in course_code else "P02"
        else:
            continue
        course_name = row[1]  # Assuming CourseName is the second column
        course_credits = row[2]  # Assuming CourseCredits is the third column
        course_description= row[3]
        deleted = row[4]  # Assuming Deleted is the fourth column

        # Append the processed data to the Course Data sheet
        course_data_sheet.append([course_code, program_id, course_name, course_credits, deleted, course_description])

    # Save the updated data.xlsx workbook
    data_workbook.save('data.xlsx')

    print("Course data processed and saved successfully.")

# Call the function to process course data
process_course_data()
