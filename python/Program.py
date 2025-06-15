import openpyxl
from openpyxl.styles import Font, Alignment

# Create a new workbook
wb = openpyxl.Workbook()

# Create a worksheet
ws = wb.active
ws.title = "Program Data"

# Set the column headings
column_headings = ["Program ID", "Program Short Name",
                   "Program Full Name", "Program Duration",
                   "Program Description"]
for col_num, heading in enumerate(column_headings, 1):
    ws.cell(row=1, column=col_num, value=heading)

# Set font and alignment for column headings
font = Font(bold=True)
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for cell in ws[1]:
    cell.font = font
    cell.alignment = alignment

# Insert the data instances
data = [
    ['P01', 'PGDSD', 'Postgraduate Diploma in Software Development', 1, 'This program is designed to enable students either to embark on a career in the information technology (IT) industry or continue their journey in academia towards a Master’s Degree.'],
    ['P02', 'B.E.', 'Bachelor’s Degree Programmes', 5, 'The bachelor’s degree programmes at MIIT have been designed keeping in mind the matriculation system of school education in Myanmar. Students in Myanmar complete the Matriculation Examination in Standard X'],
    ['P03', 'M.E.', 'Master’s Degree Programmes', 3, 'The master’s degree programmes at MIIT are full-time degree programmes conducted at the campus of MIIT in Mandalay, Myanmar. These programmes are designed to be flexible, modern, and rigorous, for a career in the IT industry.']
]

# Insert the data into the worksheet
for row_num, row_data in enumerate(data, 2):
    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=value)

# Adjust column widths
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

# Save the workbook
wb.save('data.xlsx')

print("Data instances for the Program table generated successfully in data.xlsx")
