import openpyxl

# Data for Grade and Grade Points
grades = ['A', 'A-', 'B+', 'B', 'B-', 'C+', 'C', 'D', 'F', 'S', 'X', 'W', 'I', 'RC', 'SNO', 'DNR']
grade_points = ['4.0', '3.7', '3.4', '3.0', '2.7', '2.4', '2.0', '1.0', '0.0', '0', '0', '0', '0', '0', '0', '0']

def generate_grade_point_data():
    # Load or create the workbook
    try:
        workbook = openpyxl.load_workbook('data.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Create a new sheet or get existing
    if "GradePoint Data" in workbook.sheetnames:
        grade_sheet = workbook["GradePoint Data"]
        grade_sheet.delete_rows(2, grade_sheet.max_row - 1)  # Clear existing data if any
    else:
        grade_sheet = workbook.create_sheet("GradePoint Data")

    # Add headings
    grade_sheet.append(["GradePointID", "Grade", "GradePointValue"])

    # Populate the sheet with data
    for index, (grade, point) in enumerate(zip(grades, grade_points), start=1):
        grade_point_id = f"GP{index:03d}"
        grade_sheet.append([grade_point_id, grade, point])

    # Save the workbook
    workbook.save('data.xlsx')
    print("Grade point data generated successfully.")

# Call the function
generate_grade_point_data()
