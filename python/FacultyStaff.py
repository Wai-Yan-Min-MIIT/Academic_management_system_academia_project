import openpyxl

# Load the SampleData2.xlsx workbook
sample_data_workbook = openpyxl.load_workbook('SampleData2.xlsx')
faculty_sheet = sample_data_workbook['Faculty']

# Load the data.xlsx workbook, or create it if it doesn't exist
try:
    data_workbook = openpyxl.load_workbook('data.xlsx')
except FileNotFoundError:
    data_workbook = openpyxl.Workbook()

# Create a new sheet for Faculty in data.xlsx
faculty_data_sheet = data_workbook.create_sheet("Faculty")

# Define column headings for the new sheet
headings = ["FacultyStaffID", "UserID", "Salutation", "FacultyStaffName", "ShortName", "Designation", "Department", "NRC", "Phone", "Address"]
faculty_data_sheet.append(headings)

# Process each row in the Faculty sheet from SampleData2.xlsx
for row in faculty_sheet.iter_rows(min_row=2, values_only=True):
    user_id, faculty_id, salutation, designation, faculty_name, short_name, \
    department, phone, nrc, address = row

    # Append the processed data to the Faculty Data sheet
    faculty_data_sheet.append([
        faculty_id, user_id, salutation, faculty_name, short_name, 
        designation, department, nrc, phone, address
    ])

# Save the updated data.xlsx workbook
data_workbook.save('data.xlsx')

print("Faculty data processed and saved successfully.")
