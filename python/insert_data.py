# insert_data.py
import os
import pandas as pd
import openpyxl

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "project_academia.settings")
import django

django.setup()

from academia.models import (
    Program, 
    Discipline, 
    Batch, 
    ProgramBatchDiscipline, 
    Course,
    BaseChart,
    OfferedCourses,
    AcademicYear,
    DisciplineCourse,
    GradePoint,
    MIITRole,
    MIITUsers,
    Student,
    FacultyStaff,
    Semester,
    MIITUserRole,
    SemesterGrading,
    CourseGrading,
    Project
)


def insert_data_program():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook["Program Data"]

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        program = Program(
            ProgramID=row[0],
            ProgramShortName=row[1],
            ProgramFullName=row[2],
            ProgramDuration=row[3],
            ProgramDescription=row[4]
        )
        program.save()
        print(f"Inserted Program: {program.ProgramID}")
    print("Inserted into Program table\n")


def insert_data_discipline():
    workbook = openpyxl.load_workbook('data.xlsx')

    # Insert data into Discipline table
    discipline_worksheet = workbook['Discipline Data']
    for row in discipline_worksheet.iter_rows(min_row=2, values_only=True):
        discipline = Discipline(
            DisciplineID=row[0],
            DisciplineFullName=row[1],
            DisciplineShortName=row[2],
            DisciplineDescription=row[3]
        )
        discipline.save()
        print(f"Inserted Discipline: {discipline.DisciplineID}")
    print("Inserted into Discipline table\n")


def insert_batch_data():
    workbook = openpyxl.load_workbook('data.xlsx')

    # Insert data into Batch table
    batch_worksheet = workbook['Batch Data']
    for row in batch_worksheet.iter_rows(min_row=2, values_only=True):
        batch = Batch(
            BatchID=row[0],
            BatchYear=row[1],
            TotalStudent=row[2],
            BatchDescription=row[3]
        )
        batch.save()
        print(f"Inserted Batch: {batch.BatchID}")
    print("Inserted into Batch table\n")


def insert_pbd_data():

    workbook = openpyxl.load_workbook('data.xlsx')

    pbd_worksheet = workbook['ProgramBatchDiscipline Data']

    for row in pbd_worksheet.iter_rows(min_row=2, values_only=True):

        pbd_id, program_id, batch_id, discipline_id = row

        program = Program.objects.get(ProgramID=program_id)
        batch = Batch.objects.get(BatchID=batch_id)
        discipline = Discipline.objects.get(DisciplineID=discipline_id)

        pbd = ProgramBatchDiscipline(
            PBD_ID=pbd_id,
            ProgramID=program, # pass Program object
            BatchID=batch,  
            DisciplineID=discipline
            )

        pbd.save()

        print(f"Inserted ProgramBatchDiscipline: {pbd.PBD_ID}")

    print("\nInserted into ProgramBatchDiscipline table")


def insert_course_data():
    workbook = openpyxl.load_workbook('data.xlsx')
    course_worksheet = workbook['Course Data']

    for row in course_worksheet.iter_rows(min_row=2, values_only=True):

        course_code, program_id, course_name, course_credits, deleted, course_description = row

        program = Program.objects.get(ProgramID=program_id)

        course = Course(
        CourseCode=course_code,
        ProgramID=program,
        CourseName=course_name,
        CourseCredits=course_credits,
        Deleted=deleted,
        CourseDescription=course_description
        )

        course.save()

        print(f"Inserted Course: {course.CourseCode}")

    print("\nInserted into Course table")


def insert_basechart_data():
    workbook = openpyxl.load_workbook('data.xlsx')

    basechart_worksheet = workbook['BaseChart']

    for row in basechart_worksheet.iter_rows(min_row=2, values_only=True):

        basechart_id, created_date, modified_date, program_id, batch_id, discipline_id = row

        basechart = BaseChart(
        BaseChartID = basechart_id,
        CreatedDate = created_date,
        ModifiedDate = modified_date,
        ProgramID = Program.objects.get(ProgramID=program_id),
        BatchID = Batch.objects.get(BatchID=batch_id),
        DisciplineID = Discipline.objects.get(DisciplineID=discipline_id)
        )

        basechart.save()

        print(f"Inserted BaseChart: {basechart.BaseChartID}")

    print("\nInserted into BaseChart table")

def insert_offeredcourses_data():

    workbook = openpyxl.load_workbook('data.xlsx')

    offeredcourses_worksheet = workbook['OfferedCourses']
    elective_list = []

    for row in offeredcourses_worksheet.iter_rows(min_row=2, values_only=True):

        basechart_id = row[0]
        course_code = row[1]  
        year_number = row[2]
        semester_number = row[3]

        # Get BaseChart and Course objects

        basechart = BaseChart.objects.get(BaseChartID=basechart_id) 
        try:
            course = Course.objects.get(CourseCode=course_code)

        except:
            elective_list.append(course_code)
            course = Course.objects.get(CourseCode='ELECTIVE')
            print(f'{course} INSERTED!')

        offeredcourse = OfferedCourses(
        BaseChartID=basechart,
        CourseCode=course,
        YearNumber=year_number,
        SemesterNumber=semester_number
        )

        offeredcourse.save()

        print(f"Inserted OfferedCourse: {offeredcourse}")

    print("\nInserted into OfferedCourses table")

    for i in elective_list:
        print(i)

    print(len(elective_list))


def insert_academic_year_data():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook['AcademicYear Data']

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        ay_id, start_date, end_date, create_date = row

        academic_year = AcademicYear(
        AY_ID=ay_id,
        AYStartDate=start_date,  
        AYEndDate=end_date,
        AYCreateDate=create_date
        )

        academic_year.save()

    print("Inserted data into AcademicYear table")

def insert_disciplinecourse_data():

    workbook = openpyxl.load_workbook('SampleData3.xlsx')
    worksheet = workbook['DisciplineCourse']
    counter = 0
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        counter += 1
        course_code, discipline_id, course_type = row
        print(discipline_id, course_code, course_type)
        print(f'Line : {counter}')
        if not course_code:
            continue

        discipline = Discipline.objects.get(DisciplineID=discipline_id)
        course = Course.objects.get(CourseCode=course_code)

        # Check if record already exists
        if DisciplineCourse.objects.filter(DisciplineID=discipline, CourseCode=course, CourseType=course_type).exists():
            print(f"DisciplineCourse {discipline_id} {course_code} already exists")
            continue

        disciplinecourse = DisciplineCourse(
        DisciplineID=discipline,
        CourseCode=course,
        CourseType=course_type   
        )
        disciplinecourse.save()

    print("DisciplineCourse data inserted!")


def insert_grade_point_data():
    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook['GradePoint Data']

    for row in sheet.iter_rows(min_row=2, values_only=True):
        grade_point_id, grade, grade_point_value = row
        GradePoint.objects.get_or_create(
            GradePointID=grade_point_id,
            Grade=grade,
            GradePointValue=grade_point_value
        )
        print(f"Inserted GradePoint: {grade_point_id}")

    print("All GradePoint data inserted successfully.")


def insert_miitrole_data():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook['MIITRole']

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        role_id, role_description = row

        role = MIITRole(
        RoleID=role_id,
        RoleDescription=role_description
        )

        role.save()

    print("MIITRole data inserted!")


def insert_miitusers_data():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook['MIITUsers Data']

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        user_id, email, password, status = row

        user = MIITUsers(
        UserID=user_id,
        Email=email,
        UserPasswordKey=password,
        UserStatus=status
        )

        user.save()

    print("MIITUsers data inserted!")


def insert_student_data():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook['Student']

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        student_id, user_id, discipline_id, program_id, batch_id, \
        student_name, salutation, section_name, roll_number, \
        student_nrc, student_phone, student_dob, acb_status, \
        guardian_name, guardian_nrc, guardian_phone, address = row

        # Get related foreign key instances
        try:
            user = MIITUsers.objects.get(UserID=user_id)
            program = Program.objects.get(ProgramID=program_id)
            batch = Batch.objects.get(BatchID=batch_id)
            discipline = Discipline.objects.get(DisciplineID=discipline_id)

        except:
            print(f'{user_id} does not exist.')
            continue

        # Create a new Student instance and save to the database
        Student.objects.get_or_create(
            StudentID=student_id,
            UserID=user,
            DisciplineID=discipline,
            ProgramID=program,
            BatchID=batch,
            StudentName=student_name,
            Salutation=salutation,
            SectionName=section_name,
            RollNumber=roll_number,
            StudentNRC=student_nrc or '',
            StudentPhone=student_phone or '',
            StudentDOB=student_dob,
            ACBStatus=acb_status,
            GuardianName=guardian_name or '',
            GuardianNRC=guardian_nrc or '',
            GuardianPhone=guardian_phone or '',
            Address=address or ''
        )

    print("Inserted data into Student table")


def insert_faculty_data():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook['Faculty']

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        faculty_id, user_id, salutation, faculty_name, short_name, \
        designation, department, nrc, phone, address = row

        # Get the MIITUsers instance
        try:
            user = MIITUsers.objects.get(UserID=user_id)
        except MIITUsers.DoesNotExist:
            print(f"UserID {user_id} not found in MIITUsers table.")
            continue

        # Create a new FacultyStaff instance and save to the database
        FacultyStaff.objects.get_or_create(
            FacultyStaffID=faculty_id,
            UserID=user,
            Salutation=salutation,
            FacultyStaffName=faculty_name,
            ShortName=short_name,
            Designation=designation,
            Department=department,
            NRC=nrc or '',
            Phone=phone or '',
            Address=address or ''
        )

    print("Inserted data into FacultyStaff table")


def insert_semester_data():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook['Semester']

    for row in worksheet.iter_rows(min_row=2, values_only=True):

        semester_id, ay_id, program_id, status, start_date, end_date, desc = row

        academic_year = AcademicYear.objects.get(AY_ID=ay_id)
        program = Program.objects.get(ProgramID=program_id)

        semester = Semester(
        SemesterID=semester_id,
        AY_ID=academic_year,
        ProgramID=program,
        SemesterStatus=status,
        SemesterStartDate=start_date,
        SemesterEndDate=end_date,
        SemesterDescription=desc
        )

        semester.save()

    print("Semester data inserted!")


def insert_miituserrole_data():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook['MIITUserRole']

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        user_id, role_id = row
        try:
            user = MIITUsers.objects.get(UserID=user_id)
            role = MIITRole.objects.get(RoleID=role_id)
        except:
            print(f'UserID - {user_id} does not exist.')
            continue

        miituserrole = MIITUserRole(
        UserID=user,
        RoleID=role
        )
        miituserrole.save()
    print("MIITUserRole data inserted!")


def insert_semestergrading_data():
    semester_grading_df = pd.read_excel('data.xlsx', sheet_name='SemesterGrading Data')

    for index, row in semester_grading_df.iterrows():
        
        semester_grading_id = row['SemesterGradingID']
        student_id = row['StudentID']
        semester_id = row['SemesterID']
        semester_number = row['SemesterNumber']
        deleted = row['Deleted']

        try:
            student = Student.objects.get(StudentID=student_id)  
            semester = Semester.objects.get(SemesterID=semester_id)

        except:
            print(f'Error fetching student {student_id} or semester {semester_id}')
            continue

        semester_grading = SemesterGrading(
        SemesterGradingID=semester_grading_id,
        StudentID=student,
        SemesterID=semester,
        SemesterNumber=semester_number,
        Deleted=deleted
        )
        semester_grading.save()
    
    print('SemesterGrading data inserted!')


def insert_coursegrading_data():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook['CourseGrading']

    for row in worksheet.iter_rows(min_row=2, values_only=True):

        semester_grading_id = row[0]
        course_code = row[1] 
        gradepoint_id = row[2]

        try:
            semester_grading = SemesterGrading.objects.get(SemesterGradingID=semester_grading_id)
            course = Course.objects.get(CourseCode=course_code)
            gradepoint = GradePoint.objects.get(GradePointID=gradepoint_id)

        except Exception as e:
            print(f"Error: {e}")
            continue

        course_grading = CourseGrading(
        SemesterGradingID=semester_grading,
        CourseCode=course,
        GradePointID=gradepoint
        )
        course_grading.save()

    print("CourseGrading data inserted!")


def insert_project_data():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook['Project Data']

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        
        project_code, faculty_id, semester_id, title, credits, num_students, summary, deleted, project_type, remarks = row

        try:
            faculty = FacultyStaff.objects.get(FacultyStaffID=faculty_id)
            semester = Semester.objects.get(SemesterID=semester_id)

        except Exception as e:
            print(f"Error: {e}")
            continue

        project = Project(
        ProjectNumber=project_code,
        FacultyStaffID=faculty,
        SemesterID=semester,
        ProjectTitle=title,
        ProjectCredits=credits,
        NumbStudents=num_students,
        ProjectSummary=summary,
        Deleted=deleted,
        ProjectType=project_type,
        Remarks=remarks
        )
        project.save()

    print('Project data inserted!')



if __name__ == "__main__":
    # insert_data_program()
    # insert_data_discipline()
    # insert_batch_data()
    # insert_pbd_data()
    # insert_course_data()
    insert_basechart_data()
    # insert_offeredcourses_data()
    # insert_academic_year_data()
    # insert_disciplinecourse_data()
    # insert_grade_point_data()
    # insert_miitrole_data()
    # insert_miitusers_data()
    # insert_student_data()
    # insert_faculty_data()
    # insert_semester_data()
    # insert_miituserrole_data()
    # insert_semestergrading_data()
    # insert_coursegrading_data()
    # insert_projectgrading_data()
    # insert_project_data()
    # insert_elective_data()
